import os,re,xlrd,time
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import numpy as np
import shutil
from sklearn.feature_extraction.text import TfidfVectorizer
import win32com.client as win32
import shxy_spe
import warnings
warnings.filterwarnings('ignore')


'''
文件格式转换：xls,csv,转换为xlsx
'''

def save_as_xlsx(file_path):
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(file_path)
    wb.SaveAs(file_path + 'x', FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(file_path)

def csv_save_as_xlsx(file_path):
    fname=file_path.split(sep='\\')[-1]
    file=fname[:-5]
    dirpath,ex=os.path.splitext(file_path)
    try:
        df=pd.read_csv(file_path,encoding='gbk')
        df.to_excel(f'{dirpath}.xlsx',index=False)
    except:
        df=pd.read_csv(file_path,encoding='utf-8')
        df.to_excel(f'{dirpath}.xlsx',index=False)
    os.remove(file_path)

def file_conversion(data_path):
    print('xy 开始转换文件格式，请稍后...')
    print('='*200)
    new_data_path=data_path+'处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            if file_path.endswith('.csv'):
                try:
                    csv_save_as_xlsx(file_path)#把csv文件转换为xlsx文件
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'文件转换失败->| {flaiure_path}')
                    except:
                        continue

            elif file_path.endswith('.xls'):
                try:
                    save_as_xlsx(file_path)#xls文件转换为xlsx文件
                except:
                    floder_path='\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.XLS'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path,p+'.xls')
                    file_path=p+'.xls'
                    save_as_xlsx(file_path)#xls文件转换为xlsx文件
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.XLSX'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path, p + '.xlsx')
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.file'):
                os.remove(file_path)
            elif file_path.endswith('.css'):
                os.remove(file_path)
            elif file_path.endswith('.htm'):
                os.remove(file_path)
            elif file_path.endswith('.xml'):
                os.remove(file_path)
            elif file_path.split('\\')[-1]=='.DS_Store':
                os.remove(file_path)
            elif file_path.endswith('.jpg') or file_path.endswith('.png'):
                print(f'图片文件->|{file_path}')
            else:
                continue
#检查是否转换成功
def check_conversion(data_path):
    new_data_path=data_path+'处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            try:
                df=pd.read_excel(file_path,dtype='object')
            except:
                floder_path = '\\'.join(file_path.split('\\')[:-1])
                flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                try:
                    shutil.move(floder_path, flaiure_path)
                    print(f'文件转换失败->| {flaiure_path}')
                except:
                    continue


def cheack_file(data_path):
    liste=[]
    new_data_path=data_path+'处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            p, e = os.path.splitext(file_path)
            liste.append(e)
            if e!='.xlsx':
                print(file_path)
    liste=list(set(liste))
    if liste[0]=='.xlsx':
        # print('='*120)
        print('文件格式转换完成...')
        print('='*200)
        # print('='*120)
    else:
        print('【文件格式转换未完成，请手动转换】')
        print('='*120)
        print(liste)

#上药:文件夹规整:删除空文件夹
def clean_folder2(data_path):
    if os.path.isdir(data_path):
        for i in os.listdir(data_path):
            clean_folder2(os.path.join(data_path,i))
    try:
        if not os.listdir(data_path):
            os.rmdir(data_path)
    except:
        pass

#删除空文件夹
def del_empty_floder(data_path):
    new_data_path=data_path+'处理成功'
    if os.path.isdir(new_data_path):
        for i in os.listdir(new_data_path):
            clean_folder2(os.path.join(new_data_path,i))
    try:
        if not os.listdir(new_data_path):
            os.rmdir(new_data_path)
    except:
        pass

#删除空文件夹
def del_empty_floder2(new_data_path):
    # new_data_path=data_path+'处理成功'
    if os.path.isdir(new_data_path):
        for i in os.listdir(new_data_path):
            clean_folder2(os.path.join(new_data_path,i))
    try:
        if not os.listdir(new_data_path):
            os.rmdir(new_data_path)
    except:
        pass
def shxy_sta(data_path):
    print('正在识别经销商，请稍后...')
    print('='*200)
    new_data_path = data_path + '处理成功'
    clean_path = r"C:\Users\GDY\Desktop\Work\工作\面试相关\20221121-迈睿中国\CleanSystem\data\xy\xy模糊识别库.xlsx"
    df = pd.read_excel(clean_path, usecols=['经销商文件名', '文件夹名称'])
    i = 0
    for pro in os.listdir(data_path):
        for dealer in os.listdir(f"{data_path}\\{pro}"):
            if not os.path.exists(new_data_path):
                os.makedirs(new_data_path)
            old_path = os.path.join(f"{data_path}\\{pro}", dealer)
            new_path = os.path.join(new_data_path, dealer)
            shutil.copytree(old_path, new_path)  # 复制文件夹
            pattern = re.compile('[\u4e00-\u9fa5].*[司站队部房店肃心行院]')
            result = pattern.findall(dealer)  # 清洗文件名
            if len(result) == 0:
                # folder_num=len(data_path.split('\\'))
                # new_path.split('\\')[folder_num:]
                flaiure_path= data_path + '处理失败' + '\\' + '经销商识别错误' + '\\' + '\\'.join(new_path.split('\\')[len(data_path.split('\\')):])
                shutil.move(new_path,flaiure_path)
                print(f'文件识别失败->| {flaiure_path}')
                print('-'*200)
            else:
                folder_name = df.loc[df[df['经销商文件名'] == result[0]].index, '文件夹名称'].to_list()
                if folder_name == []:
                    flaiure_path = data_path + '处理失败' + '\\' + '未识别经销商' + '\\' + '\\'.join(new_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(new_path, flaiure_path)
                    print(f'文件识别失败->| {flaiure_path}')
                    print('-' * 200)
                    # msg_tiltle = '_'.join(old_path.split(sep='\\')[-2:])
                    # wsb_path = os.path.join(new_data_path, f'{msg_tiltle}_SHXY_未识别')
                    # os.rename(new_path, wsb_path)
                    # wsb_floder = data_path + '_未识别'
                    # if not os.path.exists(wsb_floder):
                    #     os.makedirs(wsb_floder)
                    # new_wsb_path = os.path.join(wsb_floder, f'{msg_tiltle}_SHXY_未识别')
                    # shutil.move(wsb_path, new_wsb_path)
                else:
                    folder_name_path = os.path.join(new_data_path, folder_name[0])
                    folder_name_path1 = os.path.join(new_data_path, f'{folder_name[0]}1')
                    try:
                        os.rename(new_path, folder_name_path)
                    except:

                        os.rename(new_path, folder_name_path1)
                        # print('重名文件，已经重新命名...')
                    i += 1
    print(f'{i}家经销商，识别完成...')
    print('=' * 200)


# 重命名文件
def rename_file(data_path):
    print('正在重命名文件，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    # msg = str(input('请输入经销商级别->| '))
    if data_path.split('\\')[-2]=='接收文件二级商':
        msg=2
    elif data_path.split('\\')[-2]=='接收文件一级商':
        msg = 1
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            pattern = re.compile('\d{7,8}|SAL|PUR|INV')  # 提取数字 日期
            result = pattern.findall(fname)
            # file_path=print(os.path.join(dirpath,fname))
            # print(result)
            # print(file_path)
            if len(result)<1:
                flaiure_path = data_path + '处理失败' + '\\' + '流向识别失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                shutil.move(dirpath, flaiure_path)
                print(f'流向识别失败->| {flaiure_path}')
                print('-' * 200)
            else:
                if result[0] == 'INV':
                    inv_name = f"SHXY_{result[0]}_CUR_{result[1]}ZC{msg}.xlsx"  # 构造库存新文件名
                    old_file_path = os.path.join(dirpath, fname)
                    new_file_path = os.path.join(dirpath, inv_name)
                    try:
                        os.rename(old_file_path, new_file_path)
                    except:
                        continue
                elif result[0] == 'SAL' or result[0] == 'PUR':
                    file_name = f"SHXY_{result[0]}_MON_{result[1]}ZC{msg}.xlsx"  # 构造新文件名
                    old_file_path = os.path.join(dirpath, fname)
                    new_file_path = os.path.join(dirpath, file_name)
                    try:
                        os.rename(old_file_path, new_file_path)
                    except:
                        continue
                else:
                    flaiure_path = data_path + '处理失败' + '\\' + '流向识别失败' + '\\' + '\\'.join(file_path.split('\\')[len(data_path.split('\\')):-1])
                    shutil.move(dirpath, flaiure_path)
                    print(f'流向识别失败->| {flaiure_path}')
                    print('-' * 200)
    print('xy 重命名文件完成...')
    print('='*200)

def cheack_sheet(data_path):
    print('xy 正在识别多sheet文件，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            try:
                workbook = load_workbook(filename=file_path)
                sheetnames = workbook.sheetnames
            except:
                flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                if not os.path.exists(flaiure_path):
                    os.makedirs(flaiure_path)
                new_file_path = flaiure_path + '\\' + fname
                shutil.move(file_path, new_file_path)
                print(f'文件转换失败->| {flaiure_path}')
                print('-'*200)


#多sheet识别
def get_sheets(data_path):
    # print('正在识别多sheet文件，请稍后...')
    # print('=' * 200)
    new_data_path=data_path+'处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            workbook = load_workbook(filename=file_path)
            sheetnames = workbook.sheetnames
            sheet_num=0
            if len(sheetnames) > 1:
                for i in sheetnames:
                    df = pd.read_excel(file_path, dtype='object', sheetname=i)
                    if df.shape[0] > 0:
                        sheet_num+=1
                    else:
                        continue
                        # new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + i + '.xlsx'
                        # df.to_excel(new_file_path, index=False)
                # os.remove(file_path)
                if sheet_num>1:
                    floder_name = data_path + '处理失败' + '\\' + '多sheet' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                    # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：多sheet->|{file_path}')
                    print('-' * 200)

#多sheet识别 文件处理

def split_sheet(new_data_path):
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            workbook = load_workbook(filename=file_path)
            sheetnames = workbook.sheetnames
            if len(sheetnames) > 1:
                for i in sheetnames:
                    df = pd.read_excel(file_path, dtype='object', sheet_name=i)
                    if df.shape[0] == 0:
                        continue
                    else:
                        new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + i + '.xlsx'
                        df.to_excel(new_file_path, index=False)
                os.remove(file_path)

def split_sheet_rename(new_data_path):
    msg=str(input('多Sheet处理：请输入文件名日期->| '))
    msg1=str(input('多Sheet处理：经销商级别->|'))
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # pattern=re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]')
            # result=pattern.findall(fname)[0]
            result,ex=os.path.splitext(fname)
            if '汇总' in result:
                print(f'多Sheet处理：删除Sheet->| {file_path}')
                os.remove(file_path)
                print('-'*200)
            elif '销售' in result and '汇总' not in result:
                new_file_name=dirpath+'\\'+'SHXY_SAL_MON_'+msg+'ZC'+msg1+'.xlsx'
                os.rename(file_path,new_file_name)
            elif result[0:2]=='流向' or '出库' in result or result=='销':
                new_file_name=dirpath+'\\'+'SHXY_SAL_MON_'+msg+'ZC'+msg1+'.xlsx'
                os.rename(file_path,new_file_name)
            elif '库存' in result or result=='存':
                new_file_name=dirpath+'\\'+'SHXY_INV_CUR_'+msg+'ZC'+msg1+'.xlsx'
                try:
                    os.rename(file_path,new_file_name)
                    print(file_path)
                    print('='*200)
                except:
                    pass
            elif '采购' in result or '购进' in result or '入库' in result or result=='进':
                new_file_name=dirpath+'\\'+'SHXY_PUR_MON_'+msg+'ZC'+msg1+'.xlsx'
                try:
                    os.rename(file_path,new_file_name)
                except:
                    pass

def sheets_split(data_path):
    new_data_path=data_path+'处理失败'+'\\'+'多sheet'
    if os.path.exists(new_data_path):
        split_sheet(new_data_path)
        split_sheet_rename(new_data_path)
        for dirpath, dirname, filenames in os.walk(new_data_path):
            for fname in filenames:
                file_path = os.path.join(dirpath, fname)
                if fname[5:8]=='SAL':
                    floder_name=file_path.split('\\')[-2]
                    new_file_path=data_path+'处理成功'+'\\'+floder_name+'\\'+fname
                    try:
                        shutil.move(file_path,new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-'*200)
                elif fname[5:8]=='INV':
                    floder_name=file_path.split('\\')[-2]
                    new_file_path=data_path+'处理成功'+'\\'+floder_name+'\\'+fname
                    try:
                        shutil.move(file_path,new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-' * 200)
                elif fname[5:8]=='PUR':
                    floder_name=file_path.split('\\')[-2]
                    new_file_path=data_path+'处理成功'+'\\'+floder_name+'\\'+fname
                    try:
                        shutil.move(file_path,new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-' * 200)
                else:
                    print(f'多Sheet处理：处理失败->|{file_path}')
                    print('-' * 200)
    del_empty_floder2(new_data_path)


# 空文件识别；
def empty_file(data_path):
    print('正在检查空文件，请稍后...')
    print('='*200)
    new_data_path=data_path+'处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            try:
                df=pd.read_excel(file_path,dtype='object')
                if df.shape[0] == 0:
                    floder_name=data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path=floder_name+'\\'+fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                        # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：空文件->|{file_path}')
                    print('-'*200)
            except:
                print(f'empty_file 文件读取失败->|{file_path}')



#确定表头

#获取分类文本
def get_header_docs(file_path,features):
    df=pd.read_excel(file_path,dtype='object')
    column0=str()
    if df.shape[0]>0:
        for x in df.columns:
            column0=column0+','+str(x)
        docs=[]
        for i in range(df.shape[0]):
            sens=str()
            for j in df.iloc[i,:]:
                sens=sens+','+str(j)
            docs.append(sens)
        docs.insert(0,column0)
        docs.append(features)
        return docs

def TFIDF(docs):
    vectorizer=TfidfVectorizer()
    model=vectorizer.fit_transform(docs)
    tfidf=model.todense().round(6)
    # print(type(tfidf))
    return tfidf

def column_index(tfidf):
    cos_sims=[]
    row_num=len(tfidf)
    for i in range(row_num-1):
        values=tfidf[-1]
        cos_sim=(np.dot(tfidf[i],values)/(np.linalg.norm(tfidf)*np.linalg.norm(values)+1)).round(6)
        cos_sims.append(cos_sim)
    cos_max_sim=np.max(np.array(cos_sims)).round(6)
    columns_index=cos_sims.index(cos_max_sim)
    return columns_index,cos_max_sim

def get_header(data_path):
    print('正在确定表头，请稍后...')
    print('='*200)
    new_data_path=data_path+'处理成功'
    with open(r'C:\Users\guodingyu\Desktop\Guodingyu\Study\NLP\clean_data3.0\features\header_feature.txt') as f:
        header_feature = f.readlines()[0]
    header_sim_list=[]
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            # if fname[5:8]=='SAL':
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            docs=get_header_docs(file_path,header_feature)
            if df.shape[0]>1:
                tfidf=TFIDF(docs)
                header_index,cos_sim_max=column_index(tfidf)
                # print(f'表头是 {header_index} 行')
                header_sim_list.append(cos_sim_max)
                if cos_sim_max>0.00001:#cos_sim均值0.11
                    df=pd.read_excel(file_path,dtype='object',header=header_index)
                    df.to_excel(file_path,index=False)
                    pass
                else:
                    floder_name = data_path + '处理失败' + '\\' + '表头错误' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                    # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：表头错误->|{new_file_path}')
                    # print('-'*200)
                    # print(file_path)
                    # print(cos_sim_max)
                    # print(header_index)
                    # print('='*130)
        # print('='*130)
        # print(f'header_sim_list:{header_sim_list}')
        # header_sim_mean=round(np.mean(np.array(header_sim_list)),4)
        # print(f'【header_sim_mean:{header_sim_mean}】')
        # print('='*130)

#删除表头中的空字符串

def del_nullstr(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            list2=[]
            df=pd.read_excel(file_path,dtype='object')
            pattern=re.compile('[\u4e00-\u9fa5].*[A-Z]&|[\u4e00-\u9fa5].*[\u4e00-\u9fa5]')
            for i in df.columns:
                result=pattern.findall(str(i))
                if len(result)==0:
                    df.drop(i,axis=1,inplace=True)
                else:
                    list2.append(result[0])
            df.columns=list2
            df.to_excel(file_path,index=False)

# 时间格式转化
def T_date(df):
    for i in df.columns:
        if df[i].dtype == 'datetime64[ns]':
            df[i] = df[i].apply(lambda x: str(pd.to_datetime(x).date()))

#表头清洗，去除表头的空格，对‘品名-规格-产地’，‘品名规格’进行分列
def del_columns(data_path):
    new_data_path=data_path+'处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            # print(file_path)
            # fname=file_path.split(sep='\\')[-1]
            df = pd.read_excel(file_path, dtype='object')
        #     df.dropna(how='all',axis=1,inplace=True)
        #     df.columns = df.columns.map((lambda x: "".join(x.split()) if type(x) is str else x))
            for i in df.columns:
                if str(i) == '品名-规格-产地':
                    df1 = df['品名-规格-产地'].str.split('-', expand=True)
                    try:
                        df1.columns = ['品名', '规格', '产地']
                    except:
                        try:
                            df1.columns = ['品名', '规格', '产地','其他']
                        except:
                            pass
                    df2 = pd.concat([df, df1], axis=1)
                    df2.drop(['品名-规格-产地'], axis=1, inplace=True)
                    list1 = []
                    for i in df2.columns:
                        pattern = re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]|[\u4e00-\u9fa5].*|[\u4e00-\u9fa5]')
                        result = pattern.findall(i)
                        list1.append(result[0])
                    df2.columns = list1
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    # print(file_path)
                    df2.to_excel(file_path, index=False)
                elif str(i) == '品名规格':
                    df['规格'] = df['品名规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                elif str(i) == '品名/规格':
                    df['规格'] = df['品名/规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                # elif str(i).endswith('码') or str(i).endswith('编号') or i=='序号'\
                # or str(i)=='流向级别'  or 'ID' in str(i):
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)
                # elif str(i)=='序号' or str(i)=='客户ID':
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)

def split_columns(data_path):
    new_data_path=data_path+'处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            # print(file_path)
            # fname=file_path.split(sep='\\')[-1]
            df = pd.read_excel(file_path, dtype='object')
        #     df.dropna(how='all',axis=1,inplace=True)
        #     df.columns = df.columns.map((lambda x: "".join(x.split()) if type(x) is str else x))
            for i in df.columns:
                if str(i) == '品名-规格-产地':
                    df1 = df['品名-规格-产地'].str.split('-', expand=True)
                    try:
                        df1.columns = ['品名', '规格', '产地']
                    except:
                        try:
                            df1.columns = ['品名', '规格', '产地','其他']
                        except:
                            pass
                    df2 = pd.concat([df, df1], axis=1)
                    df2.drop(['品名-规格-产地'], axis=1, inplace=True)
                    list1 = []
                    for i in df2.columns:
                        pattern = re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]|[\u4e00-\u9fa5].*|[\u4e00-\u9fa5]')
                        result = pattern.findall(i)
                        list1.append(result[0])
                    df2.columns = list1
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    # print(file_path)
                    df2.to_excel(file_path, index=False)
                    print(f'xy： 品名-规格-产地 分列成功->| {file_path}')
                    print('-'*200)
                elif str(i) == '品名规格':
                    df['规格'] = df['品名规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                elif str(i) == '品名\规格':
                    df['规格'] = df['品名\规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                # elif str(i).endswith('码') or str(i).endswith('编号') or i=='序号'\
                # or str(i)=='流向级别'  or 'ID' in str(i):
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)
                # elif str(i)=='序号' or str(i)=='客户ID':
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)

#数据清洗  剔除‘编号’，‘码’,‘生产日期’，‘失效日期’之类的干扰项
# def clean_date(file_path):
#     df=pd.read_excel(file_path,dtype='object')
#     try:
#         df.dropna(how='all',axis=1,inplace=True)
#     except:
#         pass
#     for i in df.columns:
#         if str(i).endswith('码') or str(i).endswith('编号') or i=='生产日期' or i=='失效日期' or i=='序号'\
#         or i=='流向级别' or i=='客户ID' or 'ID' in i:
#             df.drop(str(i),axis=1,inplace=True)
#             df.to_excel(file_path,index=False)

#'金华市东阳医药药材有限公司':'','福建省雷允上医药有限公司':'','山西九州通医药有限责任公司':'','国药控股吉林省有限公司': '销售流向：原因待核实',
#'四川省本草堂药业有限公司': '销售流向：客户名称不易区分','邢台天宇医药有限公司': '销售流向：原因待核实','陕西省天士力医药有限公司': '销售流向：客户名称不易区分',
# '江苏宏康医药有限责任公司': '销售流向：表头错误','山东康诺盛世医药有限公司': '销售流向：原因待核实','上海市医药股份有限公司黄山华氏有限公司': '销售流向：原因待核实',
#'西安市新龙药业有限公司': '销售流向：原因待核实','十堰市君琪安药业有限公司': '销售流向：原因待核实',
def get_spe_sal(data_path):
    print('xy 正在获取黑名单经销商，请稍后...')
    print('=' * 200)
    # spe_name = {'十堰市君琪安药业有限公司': '销售流向：原因待核实',
    #             '山东康诺盛世医药有限公司': '销售流向：原因待核实',
    #             '上海市医药股份有限公司黄山华氏有限公司': '销售流向：原因待核实',
    #             '贵州科开医药股份有限公司': '销售流向：原因待核实',
    #             '西安市新龙药业有限公司': '销售流向：原因待核实',
    #             '黑龙江省大众平安医药连锁有限公司': '销售流向：原因待核实',
    #             '安徽省宿州市经济技术开发区华康医药集团有限公司': '销售流向：表头缺失',
    #             '江苏宏康医药有限责任公司': '销售流向：表头错误',
    #             '邢台天宇医药有限公司': '销售流向：原因待核实',
    #             '广西华泰药业有限公司': '销售流向：产品名称 不全',
    #             '汕头医药采购供应站': '销售流向：原因待核实',
    #             '陕西省天士力医药有限公司': '销售流向：客户名称不易区分',
    #             '福建省中源医药有限公司': '销售流向：客户名称不易区分',
    #             '国药控股吉林省有限公司': '销售流向：原因待核实',
    #             '四川省本草堂药业有限公司': '销售流向：客户名称不易区分',
    #             '重庆医药萍乡有限公司': '销售流向：原因待核实',
    #             '国药控股内江医药有限公司': '销售流向：规格缺失',
    #             '黑龙江省国龙医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
    #             '华润衢州市医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
    #             '舟山市普陀区医药药材有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
    #             '湖南省达嘉维康医药有限公司': '销售流向：客户名称客户名称不易区分',
    #             '山西省康美徕医药有限公司': '销售流向：原始文件生产厂家缺失',
    #             '苏州礼安医药销售有限公司（华润常州医药有限公司）': '销售流向：销售日期 格式错误',
    #             '江西省五洲医药营销有限公司': '销售流向：销售日期 格式错误',
    #             '江苏吴中医药销售有限公司': '销售流向：销售日期 格式错误',
    #             '浙江长典医药有限公司':'销售流向：产品名称 规格 生产厂家 在一列',
    #             '华润安徽医药有限责任公司':'销售流向：多流向',
    #             '哈药集团世一堂百川医药商贸有限公司':'销售流向：多Sheet',
    #             '河南安仁医药有限公司': '销售流向：多Sheet',
    #             '广西壮族自治区桂玉医药有限责任公司': '销售流向：多Sheet',
    #             '徐州市淮海药业有限公司': '销售流向：多Sheet',
    #             '广西桂林海王医药有限公司': '销售流向：多Sheet',
    #             '国药控股河南省股份有限公司': '销售流向：多Sheet',
    #             '兰州市阳光医药有限公司': '销售流向：多Sheet',
    #             '湖南省时代阳光医药有限公司': '销售流向：多Sheet',
    #             '瑞康医药安徽有限公司': '销售流向：多Sheet',
    #             '安徽省亳州市医药供销有限公司':'销售流向：多Sheet',
    #             }
    spe_name={'广西桂林汇通药业有限公司': '销售流向，自采数据，可能多流向',
             '揭阳市医药有限责任公司': '销售流向，自采数据，可能多流向',
             '赣州鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '泉州市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '贵州科开医药股份有限公司': '销售流向：多Sheet，数据错列',
             '浙江长典医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
             '江苏吴中医药销售有限公司': '销售流向：销售日期 格式错误',
             '华润衢州市医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
             '厦门市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '惠州市卫康药房连锁有限公司': '销售流向，自采数据，可能多流向',
             '广西华泰药业有限公司': '销售流向：产品名称 缺失',
             '舟山市普陀区医药药材有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
             '山西省康美徕医药有限公司': '销售流向：原始文件生产厂家缺失',
             '华东医药宁波销售有限责任公司': '销售流向，自采数据，可能多流向',
             '安康市长寿医药集团药业有限公司': '销售流向，自采数据，可能多流向',
             '漳州鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '莆田鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '湖南省达嘉维康医药有限公司': '销售流向：客户名称不易区分',
             '宁波市慈溪医药药材有限公司': '销售流向，自采数据，可能多流向',
             '广西贺州市医药有限责任公司': '销售流向，自采数据，可能多流向',
             '吉林省宝仁药业有限公司': '销售流向，自采数据，可能多流向',
             '湖南省华御康医药有限公司': '销售流向，自采数据，可能多流向',
             '江西省五洲医药营销有限公司': '销售流向：销售日期 格式错误,日期格式错误',
             '黑龙江省大众平安医药连锁有限公司': '销售流向：原因待核实',
             '福州市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '湖南省康恩特医药有限公司': '销售流向，自采数据，可能多流向',
             '南京市医药合肥天星有限公司': '销售流向，自采数据，可能多流向',
             '汕头医药采购供应站': '销售流向：原因待核实',
             '重庆医药萍乡有限公司': '销售流向：原因待核实',
             '宁德市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
             '黑龙江省国龙医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
             '国药控股内江医药有限公司': '销售流向：规格缺失',
             '苏州礼安医药销售有限公司（华润常州医药有限公司）': '销售流向：销售日期 格式错误',
             '长沙市同安医药有限公司': '销售流向，自采数据，可能多流向',
             '广西壮族自治区太华医药有限公司': '销售流向，自采数据，可能多流向',
             '广西广药新时代医药有限责任公司': '销售流向，自采数据，可能多流向',
             '安徽省宿州市经济技术开发区华康医药集团有限公司': '销售流向：表头缺失',
             '瑞康医药安徽有限公司': '销售流向：多Sheet',
              }
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                dps_name = file_path.split('\\')[-2].split('_')[1]
                for i in spe_name:
                    if dps_name == i:
                        floder_name = data_path + '处理失败' + '\\' + '黑名单经销商' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        if not os.path.exists(floder_name):
                            os.makedirs(floder_name)
                        new_file_path = floder_name + '\\' + fname
                        shutil.move(file_path, new_file_path)
                        print(f'xy：黑名单经销商->|{new_file_path}')
                        print(f'原因: {spe_name[i]}')
                        print('-' * 200)
                    elif dps_name == '广西毕生医药有限公司':
                        floder_name = data_path + '处理失败' + '\\' + '黑名单经销商' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        shutil.move(dirpath, floder_name)
                        print(f'xy：黑名单经销商->|{floder_name}')
                        print('-' * 200)
                        break



# def clean_spe(data_path):
#     new_data_path = data_path + '处理成功'
#     for dirpath,dirname,filenames in os.walk(new_data_path):
#         for fname in filenames:
#             file_path=os.path.join(dirpath,fname)
#             del_nullstr(file_path)#删除表头空字符串
#             del_columns(file_path)#删除容干扰的空列

def sale_date_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path=os.path.join(dirpath,fname)
                df=pd.read_excel(file_path,dtype='object')
                T_date(df)
                for i in df.columns:
                    if i == '销售日期':
                        break
                    elif i == '日期' or i == '释放日期' or i == '出库日期' or i == '流向日期' or i == '销售时间' or i == '出入库日期' or i == '发货日期'\
                            or i == '开单日期' or i=='业务日期' or i=='制单日' or i=='单据日期' or i=='业务账时间' or i=='记账时间' or i=='交货单创建日期'\
                            or i=='业务时间' or i==' 业务日期' or i=='生效时间' or i=='生效日期' or i=='出库时间' or i=='制单日期' or i=='记帐时间' or i=='确认日期'\
                            or i=='记账日期' or i=='日 期' or i=='发生日期' or i=='开票时间' or i=='订货时间' or i=='出具发票日期' or i=='c' or i=='时间'\
                            or i=='出库确认日期' or i=='开票日期' or i=='出/入库日期' or i=='创建日期' or i=='订单时间' or i=='记帐日期' or i=='发票日期'\
                            or i=='财务审核时间' or i=='购买日期' or i=='开单日期(含时分秒' or i=='凭证年度' or i=='订单日期':
                        df.rename(columns={i: '销售日期'}, inplace=True)
                        df.to_excel(file_path, index=False)
                        break
                    # elif i == '客户ID':
                    #     df.drop('客户ID',axis=1,inplace=True)
                    #     df.to_excel(file_path, index=False)


def batch_num_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            for i in df.columns:
                if i == '批号':
                    break
                elif i == '批号信息' or i == '批号效期' or i == '商品批号/效期' or i == '生产批号' or i == '订单批号' or i=='产品批号' or i=='商品批号'\
                        or i=='批号/序列号' or i=='销售批号' or i=='药品批号' or i=='货品批号' or i=='商品来源批号':
                    df.rename(columns={i: '批号'}, inplace=True)
                    df.to_excel(file_path, index=False)

def product_name_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            for i in df.columns:
                if i == '产品名称':
                    break
                elif i == '商品名称' or  i == '药品名称' or  i == '货品名称' or i=='品名' or i=='通用名' or i=='产品通用名称' or i=='通用名称':
                    df.rename(columns={i: '产品名称'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break

def product_spe_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath,dirname,filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            for i in df.columns:
                if i == '规格':
                    break
                elif i == '商品规格' or i=='药品规格' or i=='品种规格' or i=='产品规格' or i=='货品规格' or i=='规格/型号' or i=='物料规格'\
                        or i=='规 格' or i=='规格(型号' or i=='规格型号':
                    df.rename(columns={i: '规格'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break

#xy：产品编码
def product_name_id(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '产品编码':
                    break
                elif i == '商品编码' or i == '货品明细ID' or i == '商品编码' or i == '品种编码' or i == '商品编号 * 货品ID' or i == '规格/品规ID' or i == '新商品编码' \
                        or i == '药品编码' or i == '商品编号'  or i == '产品编号' or i == '商品ID' or i == '品种号' or i == '商品编码' or i == '货品编号' or i == '货号' \
                        or i == '物料编码'  or i == '商品主编码' or i == '货品编码 / 商品编码' or i == '货品编码' or i=='药品编号' or i=='药品M码' or i=='商品代码':
                    df.rename(columns={i: '产品编码'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


#xy：客户编码
def custer_name_id(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '客户编码':
                    break
                elif i == '客户号' or i == '客户ID' or i == '送达方编号' or i == '机构编码' or i == '门店编码' or i == '客户单位编码' or i == '买方编码' \
                        or i == '销往单位编码' or i == '单位编号'  or i == '客商编码' or i == '客户编码(业务码)' or i == '客户编码(内码)' or i == '销往单位编号'\
                        or i=='客户代码'  or i=='单位编码' or i=='售达方编号' or i=='客商编号':
                    df.rename(columns={i: '客户编码'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break

if __name__ == '__main__':
    data_path = r'C:\Users\guodingyu\Desktop\工具\SHXY_CLEAN\接收文件二级商\20220922'
    time_start = time.time()  # 记录开始时间
    if not os.path.exists(data_path+'处理成功'):
        shxy_sta(data_path)  #标准化  经销商识别
    get_spe_sal(data_path)  # 黑名单经销商
    file_conversion(data_path) #文件格式转换
    check_conversion(data_path)  #检查转换是否成功
    cheack_file(data_path)  #检查转换是否成功  检查后缀名
    del_empty_floder(data_path) #删除空文件夹
    shxy_spe.shxy_spe(data_path) #特殊经销商清洗
    rename_file(data_path)  #重命名文件
    cheack_sheet(data_path)  #获取sheet检查
    get_sheets(data_path)  # 多sheet识别  移动文件
    sheets_split(data_path)  #多sheet拆分
    empty_file(data_path)  # 空文件识别  移动文件，即 经销商
    get_header(data_path)  #确定表头
    split_columns(data_path)  #品名-规格-产地 分列
    sale_date_clean(data_path)  # 销售日期修改
    del_nullstr(data_path)  #删除表头空字符串
    del_columns(data_path)  #删除容干扰的空列
    batch_num_clean(data_path)  #批号修改
    product_name_clean(data_path)  #产品名称修改
    product_spe_clean(data_path)  #规格修改
    custer_name_id(data_path)  #客户编码
    product_name_id(data_path)  #产品编码

    time_end = time.time()
    time_sum = time_end - time_start
    print(f'数据清洗完成，程序运行->| {round(time_sum, 2)}s ')



