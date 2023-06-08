import os,re,xlrd,time
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import numpy as np
import shutil
from sklearn.feature_extraction.text import TfidfVectorizer
import win32com.client as win32

import warnings
warnings.filterwarnings('ignore')

'''
文件格式转换：xls,csv,转换为xlsx
'''

def save_as_xlsx(file_path):
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(file_path)
    wb.SaveAs(file_path + 'x', FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(file_path)

def csv_save_as_xlsx(file_path):
    fname=file_path.split(sep='\\')[-1]
    file=fname[:-5]
    dirpath,ex=os.path.splitext(file_path)
    try:
        df=pd.read_csv(file_path,encoding='gbk')
        df.to_excel(f'{dirpath}.xlsx',index=False)
    except:
        df=pd.read_csv(file_path,encoding='utf-8')
        df.to_excel(f'{dirpath}.xlsx',index=False)
    os.remove(file_path)

def file_conversion(data_path):
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            if file_path.endswith('.csv'):
                try:
                    csv_save_as_xlsx(file_path)#把csv文件转换为xlsx文件
                except:
                        print(f'文件转换异常->| {file_path}')
            elif file_path.endswith('.xls'):
                try:
                    save_as_xlsx(file_path)#xls文件转换为xlsx文件
                except:
                    print(f'文件转换异常->| {file_path}')
            elif file_path.endswith('.XLS'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path,p+'.xls')
                    file_path=p+'.xls'
                    save_as_xlsx(file_path)#xls文件转换为xlsx文件
                except:
                    print(f'文件转换异常->| {file_path}')
            elif file_path.endswith('.XLSX'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path, p + '.xlsx')
                except:
                    print(f'文件转换异常->| {file_path}')
            elif file_path.endswith('.file'):
                os.remove(file_path)
            elif file_path.endswith('.files'):
                shutil.rmtree(file_path)
            elif file_path.endswith('.css'):
                os.remove(file_path)
            elif file_path.endswith('.htm'):
                os.remove(file_path)
            elif file_path.endswith('.xml'):
                os.remove(file_path)
            elif file_path.split('\\')[-1]=='.DS_Store':
                os.remove(file_path)
            elif file_path.endswith('.jpg') or file_path.endswith('.png'):
                print(f'图片文件->|{file_path}')
            else:
                continue

def cheack_file(data_path):
    liste=[]
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            p, e = os.path.splitext(file_path)
            liste.append(e)
            if e!='.xlsx':
                print(file_path)
    liste=list(set(liste))
    if liste[0]=='.xlsx':
        print('='*120)
        print('【文件格式转换完成】')
        print('='*120)
    else:
        print('【文件格式转换未完成，请手动转换】')
        print('='*120)
        print(liste)
#上药:文件夹规整:提取文件
def clean_folder1(data_path):
    f_num=len(data_path.split('\\'))
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            p,e=os.path.splitext(fname)
            file_path=os.path.join(dirpath,fname)
            new_floder_path = '\\'.join(file_path.split('\\')[:f_num + 1])+'\\'+fname
            try:
                shutil.move(file_path,new_floder_path)
            except:
                pass
#上药:文件夹规整:删除空文件夹
def clean_folder2(data_path):
    if os.path.isdir(data_path):
        for i in os.listdir(data_path):
            clean_folder2(os.path.join(data_path,i))
    try:
        if not os.listdir(data_path):
            os.rmdir(data_path)
    except:
        pass

#上药:文件夹规整:构建二级文件夹
def clean_folder3(data_path):
    f_num = len(data_path.split('\\'))
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            p, e = os.path.splitext(fname)
            file_path = os.path.join(dirpath, fname)
            new_floder_path = '\\'.join(file_path.split('\\')[:f_num + 1])+'\\'+'\\'+p
            # print(new_floder_path)
            try:
                if not os.path.exists(new_floder_path):
                    os.makedirs(new_floder_path)
                new_file_path=new_floder_path+'\\'+fname
                shutil.move(file_path,new_floder_path)
            except:
                print('='*120)
                print('【构建二级文件夹失败】')
                print(file_path)
                pass
            # print(new_file_path)

#上药:文件夹规整:构建二级文件夹
def clean_folder(data_path):
    clean_folder1(data_path)
    clean_folder2(data_path)
    clean_folder3(data_path)
    print('【文件夹规整完成】')

#上药:文件夹规整:多sheet拆分
def split_sheet(data_path):
    for dirpath, dirname, filenames in os.walk(data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            workbook = load_workbook(filename=file_path)
            sheetnames = workbook.sheetnames
            if len(sheetnames) > 1:
                for i in sheetnames:
                    df = pd.read_excel(file_path, dtype='object', sheetname=i)
                    if df.shape[0] == 0:
                        continue
                    else:
                        new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + i + '.xlsx'
                        df.to_excel(new_file_path, index=False)
                os.remove(file_path)

#上药:文件夹规整:经销商识别
def get_sender_name1(data_path):
    for dirpath, dirname, filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            pattern=re.compile('[\u4e00-\u9fa5].*[司站队部房店肃心行院药燕大益]')
            result=pattern.findall(fname)
            if len(result)==0:
                floder_name=file_path.split('\\')[-2]
                result1=pattern.findall(floder_name)
                if len(result1)==0:
                    # df=pd.read_excel(file_path,dtype='object')
                    #构建经销商名称字段列表
                    # list_sender=['组织机构','机构名称','']
                    print(file_path)
                else:
                    new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + result1[0] + '.xlsx'
                    os.rename(file_path,new_file_path)

#上药:文件夹规整:经销商识别
def get_sender_name2(data_path):
    clean_path=r'C:\Users\guodingyu\Downloads\上药销售模糊识别库.xlsx'
    clean_df=pd.read_excel(clean_path,dtype='object')
    for dirpath, dirname, filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            file_name,e=os.path.splitext(fname)
            for i in clean_df['经销商文件名']:
                if i==file_name:
                    new_file_name = clean_df.loc[clean_df[clean_df['经销商文件名'] ==file_name ].index, '经销商名称'].to_list()
                    if new_file_name == []:
                        print(f'未识别经销商->|请查看{file_path}')
                    else:
                        new_file_path='\\'.join(file_path.split('\\')[:-1])+'\\'+new_file_name[0]+'.xlsx'
                        try:
                            os.rename(file_path,new_file_path)
                        except:
                            print(file_path)
                            # print(new_file_path)
                            print('【经销商识别异常】')

#上药:文件夹规整:确定表头


#获取分类文本
def get_header_docs(file_path,features):
    df=pd.read_excel(file_path,dtype='object')
    column0=str()
    if df.shape[0]>0:
        for x in df.columns:
            column0=column0+','+str(x)
        docs=[]
        for i in range(df.shape[0]):
            sens=str()
            for j in df.iloc[i,:]:
                sens=sens+','+str(j)
            docs.append(sens)
        docs.insert(0,column0)
        docs.append(features)
        return docs

def TFIDF(docs):
    vectorizer=TfidfVectorizer()
    model=vectorizer.fit_transform(docs)
    tfidf=model.todense().round(6)
    # print(type(tfidf))
    return tfidf

def column_index(tfidf):
    cos_sims=[]
    row_num=len(tfidf)
    for i in range(row_num-1):
        values=tfidf[-1]
        cos_sim=(np.dot(tfidf[i],values)/(np.linalg.norm(tfidf)*np.linalg.norm(values)+1)).round(6)
        cos_sims.append(cos_sim)
    cos_max_sim=np.max(np.array(cos_sims)).round(6)
    columns_index=cos_sims.index(cos_max_sim)
    return columns_index,cos_max_sim

def get_header(data_path):
    with open(r'C:\Users\guodingyu\Desktop\Guodingyu\Study\NLP\clean_data3.0\features\header_feature.txt') as f:
        header_feature = f.readlines()[0]
    header_sim_list=[]
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            docs=get_header_docs(file_path,header_feature)
            if df.shape[0]>1:
                tfidf=TFIDF(docs)
                header_index,cos_sim_max=column_index(tfidf)
                # print(f'表头是 {header_index} 行')
                header_sim_list.append(cos_sim_max)
                if cos_sim_max>=0.01:#cos_sim均值0.11
                    df=pd.read_excel(file_path,dtype='object',header=header_index)
                    df.to_excel(file_path,index=False)
                    pass
                else:
                    print(file_path)
                    print(cos_sim_max)
                    print(header_index)
                    print('='*130)
    print('='*130)
    print(f'header_sim_list:{header_sim_list}')
    header_sim_mean=round(np.mean(np.array(header_sim_list)),4)
    print(f'【header_sim_mean:{header_sim_mean}】')
    print('='*130)


#删除表头中的空字符串

def del_nullstr(file_path):
    list2=[]
    df=pd.read_excel(file_path,dtype='object')
    pattern=re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]')
    for i in df.columns:
        result=pattern.findall(str(i))
        if len(result)==0:
            df.drop(i,axis=1,inplace=True)
        else:
            list2.append(result[0])
    df.columns=list2
    df.to_excel(file_path,index=False)

# 时间格式转化
def T_date(df):
    for i in df.columns:
        if df[i].dtype == 'datetime64[ns]':
            df[i] = df[i].apply(lambda x: str(pd.to_datetime(x).date()))

#表头清洗，去除表头的空格，对‘品名-规格-产地’，‘品名规格’进行分列
def del_columns(file_path):
    # print(file_path)
    fname=file_path.split(sep='\\')[-1]
    df = pd.read_excel(file_path, dtype='object')
#     df.dropna(how='all',axis=1,inplace=True)
    df.columns = df.columns.map((lambda x: "".join(x.split()) if type(x) is str else x))
    for i in df.columns:
        if i == '品名-规格-产地':
            df1 = df['品名-规格-产地'].str.split('-', expand=True)
            try:
                df1.columns = ['品名', '规格', '产地']
            except:
                try:
                    df1.columns = ['品名', '规格', '产地','其他']
                except:
                    pass
            df2 = pd.concat([df, df1], axis=1)
            df2.drop(['品名-规格-产地'], axis=1, inplace=True)
            list1 = []
            for i in df2.columns:
                pattern = re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]|[\u4e00-\u9fa5].*|[\u4e00-\u9fa5]')
                result = pattern.findall(i)
                list1.append(result[0])
            df2.columns = list1
            T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
            print(file_path)
            df2.to_excel(file_path, index=False)
        elif i == '品名规格':
            df['规格'] = df['品名规格'].str.extract('(\d.*)')
            T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
            df.to_excel(file_path, index=False)
        elif str(i).endswith('码') or str(i).endswith('编号') or i=='生产日期' or i=='失效日期' or i=='序号'\
        or i=='流向级别'  or 'ID' in str(i):
            df.drop(str(i),axis=1,inplace=True)
            df.to_excel(file_path,index=False)
        elif i=='序号':
            df.drop(str(i),axis=1,inplace=True)
            df.to_excel(file_path,index=False)

#数据清洗  剔除‘编号’，‘码’,‘生产日期’，‘失效日期’之类的干扰项
def clean_date(file_path):
    df=pd.read_excel(file_path,dtype='object')
    try:
        df.dropna(how='all',axis=1,inplace=True)
    except:
        pass
    for i in df.columns:
        if str(i).endswith('码') or str(i).endswith('编号') or i=='生产日期' or i=='失效日期' or i=='序号'\
        or i=='流向级别' or i=='客户ID' or 'ID' in i:
            df.drop(str(i),axis=1,inplace=True)
            df.to_excel(file_path,index=False)

def clean_spe(data_path):
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            del_nullstr(file_path)#删除表头空字符串
            del_columns(file_path)#删除容干扰的空列

def sale_date_clean(data_path):
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            df=pd.read_excel(file_path,dtype='object')
            for i in df.columns:
                if i=='日期':
                    df.rename(columns={'日期':'销售日期'},inplace=True)
                    df.to_excel(file_path,index=False)


def rename_filename(data_path):
    for dirpath,dirname,filenames in os.walk(data_path):
        for fname in filenames:
            file_path=os.path.join(dirpath,fname)
            patttern=re.compile('[\u4e00-\u9fa5].*[司站队部房店肃心行院药燕大益]')
            result=patttern.findall(fname)
            if len(result)==0:
                print('【经销商文件名清洗失败】')
                print(file_path)
                continue
            else:
                new_file_path='\\'.join(file_path.split('\\')[:-1])+'\\'+result[0]+'.xlsx'
                os.rename(file_path,new_file_path)
                # df=pd.read_excel(new_file_path,dtype='object')
                # df['经销商文件名']=result[0]
                # df.to_excel(new_file_path,index=False)

data_path=r'C:\Users\guodingyu\Desktop\工具\SZRL_CLEAN\分拣报告第一版\20220902'
# file_conversion(data_path) #文件格式转换
# cheack_file(data_path)  #检查转换是否成功

# clean_folder1(data_path)
# clean_folder2(data_path)
# clean_folder3(data_path)
clean_folder(data_path)  #原始文件夹整理，设三级文件夹：大区\文件名\文件

split_sheet(data_path)   #多表拆分
# get_sender_name1(data_path)
# get_sender_name2(data_path) #经销商识别清洗


# get_header(data_path)  #确定表头
# clean_spe(data_path)  #特殊表头清洗
# sale_date_clean(data_path)

# rename_filename(data_path) #添加经销商文件重命名


