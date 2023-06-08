from openpyxl import load_workbook
import pandas as pd
import numpy as np
import win32com.client as win32
import shxy_pur
import shxy_inv
import shxy_spe
import os, re, time, datetime, shutil
import jieba
from sklearn.feature_extraction.text import TfidfVectorizer
from util.report import ReportCleanData
from util.tfidf import CleanModel
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import itertools
import warnings

warnings.filterwarnings('ignore')
pd.set_option('max_rows', None)
pd.set_option('max_columns', None)
pd.set_option('expand_frame_repr', False)
pd.set_option('display.unicode.east_asian_width', True)

'''
文件格式转换：xls,csv,转换为xlsx
'''


def save_as_xlsx(file_path):
    excel = win32.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(file_path)
    wb.SaveAs(file_path + 'x', FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(file_path)


def csv_save_as_xlsx(file_path):
    fname = file_path.split(sep='\\')[-1]
    file = fname[:-5]
    dirpath, ex = os.path.splitext(file_path)
    try:
        df = pd.read_csv(file_path, encoding='gbk')
        df.to_excel(f'{dirpath}.xlsx', index=False)
    except:
        df = pd.read_csv(file_path, encoding='utf-8')
        df.to_excel(f'{dirpath}.xlsx', index=False)
    os.remove(file_path)


def file_conversion(data_path):
    print('xy 开始转换文件格式，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if file_path.endswith('.csv'):
                try:
                    csv_save_as_xlsx(file_path)  # 把csv文件转换为xlsx文件
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'xy 文件转换失败->| {flaiure_path}')
                    except:
                        continue

            elif file_path.endswith('.xls'):
                try:
                    save_as_xlsx(file_path)  # xls文件转换为xlsx文件
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'xy 文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.XLS'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path, p + '.xls')
                    file_path = p + '.xls'
                    save_as_xlsx(file_path)  # xls文件转换为xlsx文件
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'xy 文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.XLSX'):
                try:
                    p, e = os.path.splitext(file_path)
                    os.rename(file_path, p + '.xlsx')
                except:
                    floder_path = '\\'.join(file_path.split('\\')[:-1])
                    flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(floder_path, flaiure_path)
                        print(f'xy 文件转换失败->| {flaiure_path}')
                    except:
                        continue
            elif file_path.endswith('.file'):
                os.remove(file_path)
            elif file_path.endswith('.css'):
                os.remove(file_path)
            elif file_path.endswith('.htm'):
                os.remove(file_path)
            elif file_path.endswith('.xml'):
                os.remove(file_path)
            elif file_path.split('\\')[-1] == '.DS_Store':
                os.remove(file_path)
            elif file_path.endswith('.jpg') or file_path.endswith('.png'):
                print(f'xy 图片文件->|{file_path}')
            else:
                continue


# 检查是否转换成功
def check_conversion(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            try:
                df = pd.read_excel(file_path, dtype='object')
            except:
                floder_path = '\\'.join(file_path.split('\\')[:-1])
                flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                try:
                    shutil.move(floder_path, flaiure_path)
                    print(f'xy 文件转换失败->| {flaiure_path}')
                except:
                    continue


def cheack_file(data_path):
    liste = []
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            p, e = os.path.splitext(file_path)
            liste.append(e)
            if e != '.xlsx':
                print(file_path)
    liste = list(set(liste))
    if liste[0] == '.xlsx':
        # print('='*120)
        print('xy 文件格式转换完成...')
        print('=' * 200)
        # print('='*120)
    else:
        print('xy 文件格式转换未完成，请手动转换')
        print('=' * 120)
        print(liste)


# 上药:文件夹规整:删除空文件夹
def clean_folder2(data_path):
    if os.path.isdir(data_path):
        for i in os.listdir(data_path):
            clean_folder2(os.path.join(data_path, i))
    try:
        if not os.listdir(data_path):
            os.rmdir(data_path)
    except:
        pass


# 删除空文件夹
def del_empty_floder(data_path):
    new_data_path = data_path + '处理成功'
    if os.path.isdir(new_data_path):
        for i in os.listdir(new_data_path):
            clean_folder2(os.path.join(new_data_path, i))
    try:
        if not os.listdir(new_data_path):
            os.rmdir(new_data_path)
    except:
        pass


# 删除空文件夹
def del_empty_floder2(new_data_path):
    # new_data_path=data_path+'处理成功'
    if os.path.isdir(new_data_path):
        for i in os.listdir(new_data_path):
            clean_folder2(os.path.join(new_data_path, i))
    try:
        if not os.listdir(new_data_path):
            os.rmdir(new_data_path)
    except:
        pass


def shxy_sta(data_path):
    print('正在识别经销商，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    clean_path = r".\data\xy模糊识别库.xlsx"
    df = pd.read_excel(clean_path, usecols=['经销商文件名', '文件夹名称'])
    i = 0
    for pro in os.listdir(data_path):
        for dealer in os.listdir(f"{data_path}\\{pro}"):
            if not os.path.exists(new_data_path):
                os.makedirs(new_data_path)
            old_path = os.path.join(f"{data_path}\\{pro}", dealer)
            new_path = os.path.join(new_data_path, dealer)
            shutil.copytree(old_path, new_path)  # 复制文件夹
            pattern = re.compile('[\u4e00-\u9fa5].*[司站队部房店肃心行院]')
            result = pattern.findall(dealer)  # 清洗文件名
            if len(result) == 0:
                # folder_num=len(data_path.split('\\'))
                # new_path.split('\\')[folder_num:]
                flaiure_path = data_path + '处理失败' + '\\' + '经销商识别错误' + '\\' + '\\'.join(
                    new_path.split('\\')[len(data_path.split('\\')):])
                shutil.move(new_path, flaiure_path)
                print(f'文件识别失败->| {flaiure_path}')
                print('-' * 200)
            else:
                folder_name = df.loc[df[df['经销商文件名'] == result[0]].index, '文件夹名称'].to_list()
                if folder_name == []:
                    flaiure_path = data_path + '处理失败' + '\\' + '未识别经销商' + '\\' + '\\'.join(
                        new_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(new_path, flaiure_path)
                    print(f'文件识别失败->| {flaiure_path}')
                    print('-' * 200)
                    # msg_tiltle = '_'.join(old_path.split(sep='\\')[-2:])
                    # wsb_path = os.path.join(new_data_path, f'{msg_tiltle}_SHXY_未识别')
                    # os.rename(new_path, wsb_path)
                    # wsb_floder = data_path + '_未识别'
                    # if not os.path.exists(wsb_floder):
                    #     os.makedirs(wsb_floder)
                    # new_wsb_path = os.path.join(wsb_floder, f'{msg_tiltle}_SHXY_未识别')
                    # shutil.move(wsb_path, new_wsb_path)
                else:
                    folder_name_path = os.path.join(new_data_path, folder_name[0])
                    folder_name_path1 = os.path.join(new_data_path, f'{folder_name[0]}1')
                    try:
                        os.rename(new_path, folder_name_path)
                    except:

                        os.rename(new_path, folder_name_path1)
                        # print('重名文件，已经重新命名...')
                    i += 1
    print(f'{i}家经销商，识别完成...')
    print('=' * 200)


# 重命名文件
def rename_file(data_path):
    print('正在重命名文件，请稍后...')
    print('=' * 200)
    msg = 1
    new_data_path = data_path + '处理成功'
    # msg = str(input('请输入经销商级别->| '))
    if data_path.split('\\')[-2] == '接收文件二级商':
        msg = 2
    elif data_path.split('\\')[-2] == '接收文件一级商':
        msg = 1
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            pattern = re.compile('\d{7,8}|SAL|PUR|INV')  # 提取数字 日期
            result = pattern.findall(fname)
            # file_path=print(os.path.join(dirpath,fname))
            # print(result)
            # print(file_path)
            if len(result) == 1:
                flaiure_path = data_path + '处理失败' + '\\' + '流向识别失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                if not os.path.exists(flaiure_path):
                    os.makedirs(flaiure_path)
                shutil.move(dirpath, flaiure_path)
                print(f'流向识别失败->| {flaiure_path}')
                print('-' * 200)
            elif len(result) == 2:
                if result[0] == 'INV':
                    inv_name = f"SHXY_{result[0]}_CUR_{result[1]}ZC{msg}.xlsx"  # 构造库存新文件名
                    old_file_path = os.path.join(dirpath, fname)
                    new_file_path = os.path.join(dirpath, inv_name)
                    try:
                        os.rename(old_file_path, new_file_path)
                    except:
                        continue
                elif result[0] == 'SAL' or result[0] == 'PUR':
                    file_name = f"SHXY_{result[0]}_MON_{result[1]}ZC{msg}.xlsx"  # 构造新文件名
                    old_file_path = os.path.join(dirpath, fname)
                    new_file_path = os.path.join(dirpath, file_name)
                    try:
                        os.rename(old_file_path, new_file_path)
                    except:
                        continue
                else:
                    flaiure_path = data_path + '处理失败' + '\\' + '流向识别失败' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    try:
                        shutil.move(dirpath, flaiure_path)
                        print(f'流向识别失败->| {flaiure_path}')
                        print('-' * 200)
                    except:
                        continue
            else:
                flaiure_path = data_path + '处理失败' + '\\' + '流向识别失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                try:
                    shutil.move(dirpath, flaiure_path)
                    print(f'流向识别失败->| {flaiure_path}')
                    print('-' * 200)
                except:
                    pass
    print('xy 重命名文件完成...')
    print('=' * 200)


def cheack_sheet(data_path):
    print('xy 正在识别多sheet文件，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            try:
                workbook = load_workbook(filename=file_path)
                sheetnames = workbook.sheetnames
            except:
                flaiure_path = data_path + '处理失败' + '\\' + '转换失败' + '\\' + '\\'.join(
                    file_path.split('\\')[len(data_path.split('\\')):-1])
                if not os.path.exists(flaiure_path):
                    os.makedirs(flaiure_path)
                new_file_path = flaiure_path + '\\' + fname
                shutil.move(file_path, new_file_path)
                print(f'文件转换失败->| {flaiure_path}')
                print('-' * 200)


# 多sheet识别
def get_sheets(data_path):
    # print('正在识别多sheet文件，请稍后...')
    # print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            workbook = load_workbook(filename=file_path)
            sheetnames = workbook.sheetnames
            sheet_num = 0
            if len(sheetnames) > 1:
                for i in sheetnames:
                    try:
                        df = pd.read_excel(file_path, dtype='object', sheetname=i)
                        if df.shape[0] > 0:
                            sheet_num += 1
                        else:
                            continue
                    except:
                        pass
                        # new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + i + '.xlsx'
                        # df.to_excel(new_file_path, index=False)
                # os.remove(file_path)
                if sheet_num > 1:
                    floder_name = data_path + '处理失败' + '\\' + '多sheet' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                    # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：多sheet->|{file_path}')
                    print('-' * 200)


# 多sheet识别 文件处理
def split_sheet(new_data_path):
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            workbook = load_workbook(filename=file_path)
            sheetnames = workbook.sheetnames
            if len(sheetnames) > 1:
                for i in sheetnames:
                    # if i=='汇总':
                    # df = pd.read_excel(file_path, dtype='object', sheet_name=i)
                    df = pd.read_excel(file_path, dtype='object', sheet_name=i)
                    if df.shape[0] == 0:
                        continue
                    else:
                        new_file_path = '\\'.join(file_path.split('\\')[:-1]) + '\\' + i + '.xlsx'
                        df.to_excel(new_file_path, index=False)
                os.remove(file_path)


def split_sheet_rename(new_data_path):
    # msg=str(input('多Sheet处理：请输入文件名日期->| '))
    # msg1=str(input('多Sheet处理：经销商级别->|'))
    if data_path.split('\\')[-2] == '接收文件二级商':
        msg = str(2)
    elif data_path.split('\\')[-2] == '接收文件一级商':
        msg = str(1)
    current_day = datetime.date.today()
    last_month_day = datetime.date(current_day.year, current_day.month, 1) - datetime.timedelta(1)
    date_year = str(last_month_day.year)
    date_month = str(last_month_day.month)
    date_day = str(last_month_day.day)
    if len(date_month) == 1:
        date_month = '0' + date_month
    else:
        date_month = date_month
    last_month_day = date_year + date_month + date_day
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # pattern=re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]')
            # result=pattern.findall(fname)[0]
            result, ex = os.path.splitext(fname)
            if '汇总' in result:
                print(f'多Sheet处理：删除Sheet->| {file_path}')
                os.remove(file_path)
                print('-' * 200)
            elif '销售' in result and '汇总' not in result:
                new_file_name = dirpath + '\\' + 'SHXY_SAL_MON_' + last_month_day + 'ZC' + msg + '.xlsx'
                os.rename(file_path, new_file_name)
            elif result[0:2] == '流向' or '出库' in result or result == '销':
                new_file_name = dirpath + '\\' + 'SHXY_SAL_MON_' + last_month_day + 'ZC' + msg + '.xlsx'
                os.rename(file_path, new_file_name)
            elif '库存' in result or result == '存':
                new_file_name = dirpath + '\\' + 'SHXY_INV_CUR_' + last_month_day + 'ZC' + msg + '.xlsx'
                try:
                    os.rename(file_path, new_file_name)
                    print(file_path)
                    print('=' * 200)
                except:
                    pass
            elif '采购' in result or '购进' in result or '入库' in result or result == '进':
                new_file_name = dirpath + '\\' + 'SHXY_PUR_MON_' + msg + 'ZC' + last_month_day + '.xlsx'
                try:
                    os.rename(file_path, new_file_name)
                except:
                    pass


def move_file(file_path, new_file_path):
    for i in range(1, 1000):
        if not os.path.exists(new_file_path):
            shutil.move(file_path, new_file_path)
        else:
            f, ex = os.path.splitext(new_file_path)
            final_file_path = f + '_RE' + str(i) + '.xlsx'
            # os.rename()
            if not os.path.exists(final_file_path):
                shutil.move(file_path, final_file_path)
                break
            else:
                continue


def sheets_split(data_path):
    new_data_path = data_path + '处理失败' + '\\' + '多sheet'
    if os.path.exists(new_data_path):
        split_sheet(new_data_path)  # 文件拆分
        split_sheet_rename(new_data_path)  # 拆分文件处理
        for dirpath, dirname, filenames in os.walk(new_data_path):
            for fname in filenames:
                file_path = os.path.join(dirpath, fname)
                if fname[5:8] == 'SAL':
                    floder_name = file_path.split('\\')[-2]
                    new_file_path = data_path + '处理成功' + '\\' + floder_name + '\\' + fname
                    try:
                        move_file(file_path, new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-' * 200)
                elif fname[5:8] == 'INV':
                    floder_name = file_path.split('\\')[-2]
                    new_file_path = data_path + '处理成功' + '\\' + floder_name + '\\' + fname
                    try:
                        move_file(file_path, new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-' * 200)
                elif fname[5:8] == 'PUR':
                    floder_name = file_path.split('\\')[-2]
                    new_file_path = data_path + '处理成功' + '\\' + floder_name + '\\' + fname
                    try:
                        move_file(file_path, new_file_path)
                    except:
                        print(f'多Sheet处理：文件已存在->|{file_path}')
                        print('-' * 200)
                else:
                    print(f'多Sheet处理：处理失败->|{file_path}')
                    print('-' * 200)
    del_empty_floder2(new_data_path)


# 空文件识别；
def empty_file(data_path):
    print('xy： 正在检查空文件，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            try:
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] == 0:
                    floder_name = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                    # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：空文件->|{file_path}')
                    print('-' * 200)
            except:
                print(f'empty_file 文件读取失败->|{file_path}')


# 确定表头

# 获取分类文本
def get_header_docs(file_path, features):
    df = pd.read_excel(file_path, dtype='object')
    column0 = str()
    if df.shape[0] > 0:
        for x in df.columns:
            column0 = column0 + ',' + str(x)
        docs = []
        for i in range(df.shape[0]):
            sens = str()
            for j in df.iloc[i, :]:
                sens = sens + ',' + str(j)
            docs.append(sens)
        docs.insert(0, column0)
        docs.append(features)
        return docs


def TFIDF(docs):
    vectorizer = TfidfVectorizer()
    model = vectorizer.fit_transform(docs)
    tfidf = model.todense().round(6)
    # print(type(tfidf))
    return tfidf


def column_index(tfidf):
    cos_sims = []
    row_num = len(tfidf)
    for i in range(row_num - 1):
        values = tfidf[-1]
        cos_sim = (np.dot(tfidf[i], values) / (np.linalg.norm(tfidf) * np.linalg.norm(values) + 1)).round(6)
        cos_sims.append(cos_sim)
    cos_max_sim = np.max(np.array(cos_sims)).round(6)
    columns_index = cos_sims.index(cos_max_sim)
    return columns_index, cos_max_sim


def get_header(data_path):
    print('正在确定表头，请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    with open(r'.\features\shxy\header_feature.txt') as f:
        header_feature = f.readlines()[0]
    header_sim_list = []
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            # if fname[5:8]=='SAL':
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            docs = get_header_docs(file_path, header_feature)
            if df.shape[0] > 1:
                tfidf = TFIDF(docs)
                header_index, cos_sim_max = column_index(tfidf)
                # print(f'表头是 {header_index} 行')
                header_sim_list.append(cos_sim_max)
                if cos_sim_max > 0.00001:  # cos_sim均值0.11
                    df = pd.read_excel(file_path, dtype='object', header=header_index)
                    df.to_excel(file_path, index=False)
                    pass
                else:
                    floder_name = data_path + '处理失败' + '\\' + '表头错误' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    # flaiure_path = data_path + '处理失败' + '\\' + '空文件' + '\\' + '\\'.join(
                    # file_path.split('\\')[len(data_path.split('\\')):])
                    shutil.move(file_path, new_file_path)
                    print(f'xy：表头错误->|{new_file_path}')
                    # print('-'*200)
                    # print(file_path)
                    # print(cos_sim_max)
                    # print(header_index)
                    # print('='*130)
        # print('='*130)
        # print(f'header_sim_list:{header_sim_list}')
        # header_sim_mean=round(np.mean(np.array(header_sim_list)),4)
        # print(f'【header_sim_mean:{header_sim_mean}】')
        # print('='*130)


# 删除表头中的空字符串

def del_nullstr(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            list2 = []
            df = pd.read_excel(file_path, dtype='object')
            pattern = re.compile('[\u4e00-\u9fa5].*[A-Z]|[\u4e00-\u9fa5].*[\u4e00-\u9fa5]')
            for i in df.columns:
                result = pattern.findall(str(i))
                if len(result) == 0:
                    df.drop(i, axis=1, inplace=True)
                else:
                    list2.append(result[0])
            df.columns = list2
            df.to_excel(file_path, index=False)


# 时间格式转化
def T_date(df):
    for i in df.columns:
        if df[i].dtype == 'datetime64[ns]':
            df[i] = df[i].apply(lambda x: str(pd.to_datetime(x).date()))


# 表头清洗，去除表头的空格，对‘品名-规格-产地’，‘品名规格’进行分列
def del_columns(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            # fname=file_path.split(sep='\\')[-1]
            df = pd.read_excel(file_path, dtype='object')
            #     df.dropna(how='all',axis=1,inplace=True)
            #     df.columns = df.columns.map((lambda x: "".join(x.split()) if type(x) is str else x))
            for i in df.columns:
                if str(i) == '品名-规格-产地':
                    df1 = df['品名-规格-产地'].str.split('-', expand=True)
                    try:
                        df1.columns = ['品名', '规格', '产地']
                    except:
                        try:
                            df1.columns = ['品名', '规格', '产地', '其他']
                        except:
                            pass
                    df2 = pd.concat([df, df1], axis=1)
                    df2.drop(['品名-规格-产地'], axis=1, inplace=True)
                    list1 = []
                    for i in df2.columns:
                        pattern = re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]|[\u4e00-\u9fa5].*|[\u4e00-\u9fa5]')
                        result = pattern.findall(i)
                        list1.append(result[0])
                    df2.columns = list1
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    # print(file_path)
                    df2.to_excel(file_path, index=False)
                elif str(i) == '品名规格':
                    df['规格'] = df['品名规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                elif str(i) == '品名/规格':
                    df['规格'] = df['品名/规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)


def split_columns(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            # fname=file_path.split(sep='\\')[-1]
            df = pd.read_excel(file_path, dtype='object')
            #     df.dropna(how='all',axis=1,inplace=True)
            #     df.columns = df.columns.map((lambda x: "".join(x.split()) if type(x) is str else x))
            for i in df.columns:
                if str(i) == '品名-规格-产地':
                    df1 = df['品名-规格-产地'].str.split('-', expand=True)
                    try:
                        df1.columns = ['品名', '规格', '产地']
                    except:
                        try:
                            df1.columns = ['品名', '规格', '产地', '其他']
                        except:
                            pass
                    df2 = pd.concat([df, df1], axis=1)
                    df2.drop(['品名-规格-产地'], axis=1, inplace=True)
                    list1 = []
                    for i in df2.columns:
                        pattern = re.compile('[\u4e00-\u9fa5].*[\u4e00-\u9fa5]|[\u4e00-\u9fa5].*|[\u4e00-\u9fa5]')
                        result = pattern.findall(i)
                        list1.append(result[0])
                    df2.columns = list1
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    # print(file_path)
                    df2.to_excel(file_path, index=False)
                    print(f'xy： 品名-规格-产地 分列成功->| {file_path}')
                    print('-' * 200)
                elif str(i) == '品名规格':
                    df['规格'] = df['品名规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                elif str(i) == '品名\规格':
                    df['规格'] = df['品名\规格'].str.extract('(\d.*)')
                    T_date(df)  # 日期处理  年-月-日 时-分-秒  转换为 年-月-日
                    df.to_excel(file_path, index=False)
                # elif str(i).endswith('码') or str(i).endswith('编号') or i=='序号'\
                # or str(i)=='流向级别'  or 'ID' in str(i):
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)
                # elif str(i)=='序号' or str(i)=='客户ID':
                #     df.drop(str(i),axis=1,inplace=True)
                #     df.to_excel(file_path,index=False)


# 数据清洗  剔除‘编号’，‘码’,‘生产日期’，‘失效日期’之类的干扰项
# def clean_date(file_path):
#     df=pd.read_excel(file_path,dtype='object')
#     try:
#         df.dropna(how='all',axis=1,inplace=True)
#     except:
#         pass
#     for i in df.columns:
#         if str(i).endswith('码') or str(i).endswith('编号') or i=='生产日期' or i=='失效日期' or i=='序号'\
#         or i=='流向级别' or i=='客户ID' or 'ID' in i:
#             df.drop(str(i),axis=1,inplace=True)
#             df.to_excel(file_path,index=False)

# '金华市东阳医药药材有限公司':'','福建省雷允上医药有限公司':'','山西九州通医药有限责任公司':'','国药控股吉林省有限公司': '销售流向：原因待核实',
# '四川省本草堂药业有限公司': '销售流向：客户名称不易区分','邢台天宇医药有限公司': '销售流向：原因待核实','陕西省天士力医药有限公司': '销售流向：客户名称不易区分',
# '江苏宏康医药有限责任公司': '销售流向：表头错误','山东康诺盛世医药有限公司': '销售流向：原因待核实','上海市医药股份有限公司黄山华氏有限公司': '销售流向：原因待核实',
# '西安市新龙药业有限公司': '销售流向：原因待核实','十堰市君琪安药业有限公司': '销售流向：原因待核实','江西省五洲医药营销有限公司': '销售流向：销售日期 格式错误,日期格式错误',
# '广西华泰药业有限公司': '销售流向：产品名称 缺失','江苏吴中医药销售有限公司': '销售流向：销售日期 格式错误', '重庆医药萍乡有限公司': '销售流向：原因待核实',
# '山西省康美徕医药有限公司': '销售流向：原始文件生产厂家缺失',
def get_spe_sal(data_path):
    print('xy 正在获取黑名单经销商，请稍后...')
    print('=' * 200)
    spe_name = {'广西桂林汇通药业有限公司': '销售流向，自采数据，可能多流向',
                '揭阳市医药有限责任公司': '销售流向，自采数据，可能多流向',
                '赣州鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '泉州市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '贵州科开医药股份有限公司': '销售流向：多Sheet，数据错列',
                '浙江长典医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
                '华润衢州市医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
                '厦门市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '惠州市卫康药房连锁有限公司': '销售流向，自采数据，可能多流向',
                '舟山市普陀区医药药材有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
                '华东医药宁波销售有限责任公司': '销售流向，自采数据，可能多流向',
                '安康市长寿医药集团药业有限公司': '销售流向，自采数据，可能多流向',
                '漳州鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '莆田鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '湖南省达嘉维康医药有限公司': '销售流向：客户名称不易区分',
                '宁波市慈溪医药药材有限公司': '销售流向，自采数据，可能多流向',
                '广西贺州市医药有限责任公司': '销售流向，自采数据，可能多流向',
                '吉林省宝仁药业有限公司': '销售流向，自采数据，可能多流向',
                '湖南省华御康医药有限公司': '销售流向，自采数据，可能多流向',
                '黑龙江省大众平安医药连锁有限公司': '销售流向：原因待核实',
                '福州市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '湖南省康恩特医药有限公司': '销售流向，自采数据，可能多流向',
                '南京市医药合肥天星有限公司': '销售流向，自采数据，可能多流向',
                '汕头医药采购供应站': '销售流向：合并单元格，格式混乱',
                '宁德市鹭燕医药有限公司': '销售流向，自采数据，可能多流向',
                '黑龙江省国龙医药有限公司': '销售流向：产品名称 规格 生产厂家 在一列',
                '国药控股内江医药有限公司': '销售流向：规格缺失',
                '苏州礼安医药销售有限公司（华润常州医药有限公司）': '销售流向：销售日期 格式错误',
                '长沙市同安医药有限公司': '销售流向，自采数据，可能多流向',
                '广西壮族自治区太华医药有限公司': '销售流向，自采数据，可能多流向',
                '广西广药新时代医药有限责任公司': '销售流向，自采数据，可能多流向',
                '安徽省宿州市经济技术开发区华康医药集团有限公司': '销售流向：表头缺失',
                '瑞康医药安徽有限公司': '销售流向：多Sheet',
                '宜宾市众生医药有限公司(原宜宾科伦医药有限公司)': '销售流向，自采数据，可能多流向',
                '福建省雷允上医药有限公司': '销售流向，品名-规格-产地 可能分列失败',
                '安徽乐嘉医药科技有限公司': '销售流向，销售日期 可能格式错误',
                '湖州市英特药业有限公司': '销售流向，多Sheet拆分',
                '浙江省来益医药有限公司': '销售流向，数据格式混乱',
                '兰州市强生医药有限责任公司': '销售流向，销售日期数据错误',
                '国药控股河南省股份有限公司': '销售流向:一级商保留总公司数据',
                '四川省医药股份有限公司': '销售流向:一级商保留总公司数据',
                }
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                dps_name = file_path.split('\\')[-2].split('_')[1]
                for i in spe_name:
                    if dps_name == i:
                        floder_name = data_path + '处理失败' + '\\' + '黑名单经销商' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        if not os.path.exists(floder_name):
                            os.makedirs(floder_name)
                        new_file_path = floder_name + '\\' + fname
                        shutil.move(file_path, new_file_path)
                        print(f'xy：黑名单经销商->|{new_file_path}')
                        print(f'原因: {spe_name[i]}')
                        print('-' * 200)
                    elif dps_name == '广西毕生医药有限公司':
                        floder_name = data_path + '处理失败' + '\\' + '黑名单经销商' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        shutil.move(dirpath, floder_name)
                        print(f'xy：黑名单经销商->|{floder_name}')
                        print('-' * 200)
                        break


# def clean_spe(data_path):
#     new_data_path = data_path + '处理成功'
#     for dirpath,dirname,filenames in os.walk(new_data_path):
#         for fname in filenames:
#             file_path=os.path.join(dirpath,fname)
#             del_nullstr(file_path)#删除表头空字符串
#             del_columns(file_path)#删除容干扰的空列

def del_temporary(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[0] == '~':
                try:
                    os.remove(file_path)
                except Exception as e:
                    print('xy临时文件删除异常', e)
                    print('-' * 200)


def sale_date_clean1(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                T_date(df)
                if '销售日期' in list(df.columns):
                    break
                else:
                    for i in df.columns:
                        if i == '日期' or i == '出库日期 ' or i == '释放日期' or i == '出库日期' or i == '流向日期' or i == '销售时间' or i == '出入库日期' or i == '发货日期' \
                                or i == '开单日期' or i == '业务日期' or i == '制单日' or i == '单据日期' or i == '业务账时间' or i == '记账时间' or i == '交货单创建日期' \
                                or i == '业务时间' or i == ' 业务日期' or i == '生效时间' or i == '生效日期' or i == '出库时间' or i == '制单日期' or i == '记帐时间' or i == '确认日期' \
                                or i == '记账日期' or i == '日 期' or i == '发生日期' or i == '开票时间' or i == '订货时间' or i == '出具发票日期' or i == 'c' or i == '时间' \
                                or i == '出库确认日期' or i == '开票日期' or i == '出/入库日期' or i == '创建日期' or i == '订单时间' or i == '记帐日期' or i == '发票日期' \
                                or i == '财务审核时间' or i == '购买日期' or i == '开单日期(含时分秒)' or i == '凭证年度' or i == '订单日期' or i == '开单日期(含时分秒' \
                                or i == '开单时间' or i == '结算日期' or i == '销售确认日期' or i == '实际出库日期(记保管账' or i == '实际出库日期(记保管账)' or i == '开票确认时间':
                            df.rename(columns={i: '销售日期'}, inplace=True)
                            df.to_excel(file_path, index=False)
                            break
                    # elif i == '客户ID':
                    #     df.drop('客户ID',axis=1,inplace=True)
                    #     df.to_excel(file_path, index=False)


def batch_num_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '批号':
                    break
                elif i == '批号信息' or i == '批号效期' or i == '商品批号/效期' or i == '生产批号' or i == '订单批号' or i == '产品批号' or i == '商品批号' \
                        or i == '批号/序列号' or i == '销售批号' or i == '药品批号' or i == '货品批号' or i == '商品来源批号':
                    df.rename(columns={i: '批号'}, inplace=True)
                    df.to_excel(file_path, index=False)


def product_name_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '产品名称':
                    break
                elif i == '商品名称' or i == '药品名称' or i == '货品名称' or i == '品名' or i == '通用名' or i == '产品通用名称' or i == '通用名称':
                    df.rename(columns={i: '产品名称'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


def product_spe_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '规格':
                    break
                elif i == '商品规格' or i == '药品规格' or i == '品种规格' or i == '产品规格' or i == '货品规格' or i == '规格/型号' or i == '物料规格' \
                        or i == '规 格' or i == '规格(型号' or i == '规格型号' or i == '商品规格/型号':
                    df.rename(columns={i: '规格'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


# xy：数量
def product_num_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[5:8] == 'SAL':
                df = pd.read_excel(file_path, dtype='object')
                for i in df.columns:
                    if i == '数量':
                        break
                    elif i == '销售数量' or i == '出库数量' or i == '实际数量' or i == '出入库数量' or i == '出库数' or i == '商品数量' or i == '基本单位数量' \
                            or i == '流动数量' or i == '实发数量' or i == '交货数量':
                        df.rename(columns={i: '数量'}, inplace=True)
                        df.to_excel(file_path, index=False)
                        break


# xy：客户名称
def custer_name_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '客户名称':
                    break
                elif i == '往来单位名称' or i == '客户名' or i == '客商名称' or i == '售达方描述' or i == '送达方名称':
                    df.rename(columns={i: '客户名称'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


# xy：产品编码
def product_name_id(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '产品编码':
                    break
                elif i == '商品编码' or i == '货品明细ID' or i == '商品编码' or i == '品种编码' or i == '商品编号 * 货品ID' or i == '规格/品规ID' or i == '新商品编码' \
                        or i == '药品编码' or i == '商品编号' or i == '产品编号' or i == '商品ID' or i == '品种号' or i == '商品编码' or i == '货品编号' or i == '货号' \
                        or i == '物料编码' or i == '商品主编码' or i == '货品编码 / 商品编码' or i == '货品编码' or i == '药品编号' or i == '药品M码' or i == '商品代码' or i == '公司商品编码':
                    df.rename(columns={i: '产品编码'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


# xy：效期
def deter_date_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                for i in df.columns:
                    if i == '有效期':
                        df.rename(columns={'有效期': '效期'}, inplace=True)
                        # T_date(df)
                        df.to_excel(file_path, index=False)
                    elif i == '失效日期':
                        df.rename(columns={'失效日期': '效期'}, inplace=True)
                        # T_date(df)
                        df.to_excel(file_path, index=False)


def get_deter_date(data_path):
    # print('xy库存：正在提取 有效期 ,请稍后...')
    # print('='*200)
    new_data_path = data_path + '处理成功'
    mapping_custer_name = ['有效期至', '有效期', '失效日期', '有效日期', '保质日期', '保质期至', '失效期', '商品有效期', \
                           '有效期至/失效日期', '药品有效期至/医疗器械失效日期', '时效日期', '有效期限', '有效月',
                           '有效期(月)', '有效期(月', '灭菌有效期']
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0 and '效期' in list(df.columns):
                    break
                else:
                    list1 = df.columns
                    column_names = [column_name for column_name in list1 if column_name in mapping_custer_name]
                    if len(column_names) == 0:

                        pass
                    elif len(column_names) == 1 and column_names[0] != '效期':
                        df['效期'] = df.loc[:, column_names[0]]
                        #                 df.rename(columns={column_names[0]:'客户名称'},inplace=True)
                        df.to_excel(file_path, index=False)
                    elif len(column_names) > 1:
                        if '效期' not in column_names:
                            print(column_names)
                            print(file_path)
                            print('=' * 200)
                            msg = str(input('请输入 效期 字段,输入N手动处理 '))
                            if msg == 'N':
                                floder_name = data_path + '处理失败' + '\\' + '效期' + '\\' + '\\'.join(
                                    file_path.split('\\')[len(data_path.split('\\')):-1])
                                if not os.path.exists(floder_name):
                                    os.makedirs(floder_name)
                                new_file_path = floder_name + '\\' + fname
                                shutil.move(file_path, new_file_path)
                            else:
                                df.rename(columns={msg: '效期'}, inplace=True)
                                df.to_excel(file_path, index=False)


# xy：客户编码
def custer_name_id(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            df = pd.read_excel(file_path, dtype='object')
            for i in df.columns:
                if i == '客户代码':
                    break
                elif i == '客户编码' or i == '客户号' or i == '客户ID' or i == '送达方编号' or i == '机构编码' or i == '门店编码' or i == '客户单位编码' or i == '买方编码' \
                        or i == '销往单位编码' or i == '单位编号' or i == '客商编码' or i == '客户编码(业务码)' or i == '客户编码(内码)' or i == '销往单位编号' \
                        or i == '客户代码' or i == '单位编码' or i == '售达方编号' or i == '客商编号':
                    df.rename(columns={i: '客户代码'}, inplace=True)
                    df.to_excel(file_path, index=False)
                    break


'''
xy：特征工程
'''


# xy：销售日期特征
def sale_date_features():
    # start_date=str(input('xy： 请输入业务起始日期 '))
    # end_date = str(input('xy： 请输入业务终止日期 '))
    start_date = '20220801'
    end_date = '20221031'
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    date_series1 = pd.date_range(start_date, end_date).strftime('%Y-%m-%d').to_list()
    date_series2 = pd.date_range(start_date, end_date).strftime('%Y%m%d').to_list()
    date_series3 = date_series1 + date_series2
    sens_date_series = str()
    for i in date_series3:
        sens_date_series = str(i) + ' ' + sens_date_series
    sens_date = jieba.cut(sens_date_series, cut_all=False)
    sens_date = sorted(list(set(sens_date)))
    date_str = str()
    for i in sens_date:
        date_str = str(i) + ' ' + date_str
    return date_str


def tfidf(df, features):
    global columns_index, cos_sim_max
    docs = []
    for i in range(df.shape[1]):
        sens = str()
        for j in df.iloc[:, i]:
            sens = str(j) + ' ' + sens
        tokens = list(set(jieba.cut(sens, cut_all=False)))
        token = str()
        for x in tokens:
            token = x + ' ' + token
        docs.append(token)
    docs.append(features)
    vectorizer = TfidfVectorizer()
    model = vectorizer.fit_transform(docs)
    tfidf = model.todense().round(6)
    return tfidf


def column_index(tfidf):
    cos_sims = []
    for i in range(len(tfidf) - 1):
        values = tfidf[-1]
        cos_sim = (np.dot(tfidf[i], values) / (np.linalg.norm(tfidf) * np.linalg.norm(values) + 1)).round(6)
        cos_sims.append(cos_sim)
    cos_max_sim = np.max(np.array(cos_sims)).round(6)
    # print(cos_sims)
    columns_index = cos_sims.index(cos_max_sim)
    return columns_index, cos_max_sim


# xy  销售日期
def get_sale_date(data_path):
    print('xy： 正在映射 销售日期,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    list_sale_date_sim = []
    sale_date_feature = sale_date_features()
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            # mapping_sale_date(file_path)
            if fname[5:8] == 'SAL':
                df = pd.read_excel(file_path, dtype='object')
                if '销售日期' not in list(df.columns) and df.shape[0] > 0:
                    #             print(file_path)
                    #             docs=get_docs(df,sale_date_feature)
                    # print(file_path)
                    # print(docs)
                    tfidf_value = tfidf(df, sale_date_feature)
                    sale_date_index, cos_sim_max = column_index(tfidf_value)
                    list_sale_date_sim.append(cos_sim_max)
                    df = pd.read_excel(file_path, dtype='object')
                    if cos_sim_max >= 0.1:  # cos_sim均值0.1645
                        df['销售日期'] = df.iloc[:, sale_date_index]
                        df.to_excel(file_path, index=False)
                        pass
                    else:
                        floder_name = data_path + '处理失败' + '\\' + '销售日期' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        if not os.path.exists(floder_name):
                            os.makedirs(floder_name)
                        new_file_path = floder_name + '\\' + fname
                        shutil.move(file_path, new_file_path)
                        print(f'xy： 销售日期 失败->|{new_file_path}')
                        print('-' * 200)
                        # print(file_path)
                        # print(cos_sim_max)
                        # print(list(df.columns)[sale_date_index])
                        # print('='*130)

    # print('='*130)
    # print(f'list_factory_sim:{list_sale_date_sim}')
    # saledate_sim_mean=round(np.mean(np.array(list_sale_date_sim)),4)
    # print(f'【saledate_sim_mean:{saledate_sim_mean}】')
    # print('='*130)


# 深圳瑞霖：客户名称
def get_custer_name(data_path):
    print('xy： 正在映射 客户名称,请稍后...')
    print('-' * 200)
    new_data_path = data_path + '处理成功'
    mapping_custer_name = ['客户名称', '客户简称', '客户名', '销售客户', '销往单位', '客商名称', '收货客商名称', '单位全名', '医院名称',
                           '买方原名', '收货单位', '销货单位', '销售客户名称', '售达方描述', '销往单位名称', '送达方名称', '下游客户',
                           '目标业务机构', '发运至名称', '往来单位名称', '单位名', '购入客户名称', '购货单位', '相关企业名称',
                           '客户全称', '单位名称', '客商', '售达方名称', '购进单位', '流向单位', '销往地点', '相关单位名称', '商品去向',
                           '订单客户', '往来单位', '售达方', '客户单位', '客户', '销售单位名称', '销售单位', '接收方门店名称',
                           '收发货单位名称', '业务单位名称', '销售商名称', ' 供应商/客户', '商户名称', '顾客名', '采购方名称', '客户(供应商', '销 售 单 位',
                           '购货单位名称', '业务机构.1', '业务机构名称', '门店名称', '购货客户', '相关企业', '机构名称', '配送门店名称', '接收方描述', '店铺名称',
                           '下游名称', '客户/供应商']
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    list1 = df.columns
                    column_names = [column_name for column_name in list1 if column_name in mapping_custer_name]
                    if len(column_names) == 0:
                        floder_name = data_path + '处理失败' + '\\' + '客户名称' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        if not os.path.exists(floder_name):
                            os.makedirs(floder_name)
                        new_file_path = floder_name + '\\' + fname
                        shutil.move(file_path, new_file_path)

                        print(df.columns)
                        print(file_path)
                        print('-' * 200)
                        # msg=str(input('请添加【客户名称】映射字段,输入N手动处理 '))
                        # if msg == 'N':
                        #     floder_name = data_path + '处理失败' + '\\' + '客户名称' + '\\' + '\\'.join(
                        #         file_path.split('\\')[len(data_path.split('\\')):-1])
                        #     if not os.path.exists(floder_name):
                        #         os.makedirs(floder_name)
                        #     new_file_path = floder_name + '\\' + fname
                        #     shutil.move(file_path, new_file_path)
                    #         else:
                    #             df['客户名称']=df.loc[:,msg]
                    # #                 df.rename(columns={msg:'客户名称'},inplace=True)
                    #             df.to_excel(file_path,index=False)
                    elif len(column_names) == 1 and column_names[0] != '客户名称':
                        df['客户名称'] = df.loc[:, column_names[0]]
                        #                 df.rename(columns={column_names[0]:'客户名称'},inplace=True)
                        df.to_excel(file_path, index=False)
                    elif len(column_names) > 1:
                        if '客户名称' not in column_names:
                            floder_name = data_path + '处理失败' + '\\' + '客户名称' + '\\' + '\\'.join(
                                file_path.split('\\')[len(data_path.split('\\')):-1])
                            if not os.path.exists(floder_name):
                                os.makedirs(floder_name)
                            new_file_path = floder_name + '\\' + fname
                            shutil.move(file_path, new_file_path)

                            print(column_names)
                            print(file_path)
                            print('-' * 200)
                            # msg=str(input('请输入客户名称字段,输入N手动处理 '))
                            # if msg=='N':
                            #     floder_name = data_path + '处理失败' + '\\' + '客户名称' + '\\' + '\\'.join(
                            #         file_path.split('\\')[len(data_path.split('\\')):-1])
                            #     if not os.path.exists(floder_name):
                            #         os.makedirs(floder_name)
                            #     new_file_path = floder_name + '\\' + fname
                            #     shutil.move(file_path, new_file_path)
                            # else:
                            #     df.rename(columns={msg:'客户名称'},inplace=True)
                            #     df.to_excel(file_path,index=False)

                else:
                    print(f'空文件-> {file_path}')


# 深圳瑞霖：数量
def get_product_num(data_path):
    print('xy： 正在识别 数量,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    maping_num = ['数量', '实际出库数量', '销售数量', '实际出库数量', '商品数量', '出库数', '实发数量', '出库数量', '流动数量', '流向数量', '实际数量', '药品数量',
                  '开票数量',
                  '结算数量', '交货数量', '出入库数量', '基本单位数量', '订单数量', '供应数量', '出货', '实出辅数量', '原始数量', '实际发货数量',
                  '销/销退数量', '出数量', '产品数量', '商品销售数量', '发出)数量(库存', '帐本付出', '销售数量(汇总', '数量汇总', '销售/调拨数量', '数量[汇总',
                  '可销数量', '出货', '已出发票数量']
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    list1 = df.columns
                    column_names = [column_name for column_name in list1 if column_name in maping_num]
                    if len(column_names) == 0:
                        print(df.columns)
                        print(file_path)
                        print('-' * 200)
                        floder_name = data_path + '处理失败' + '\\' + '数量' + '\\' + '\\'.join(
                            file_path.split('\\')[len(data_path.split('\\')):-1])
                        if not os.path.exists(floder_name):
                            os.makedirs(floder_name)
                        new_file_path = floder_name + '\\' + fname
                        shutil.move(file_path, new_file_path)
                    elif len(column_names) == 1:
                        df.rename(columns={column_names[0]: '数量'}, inplace=True)
                        df.to_excel(file_path, index=False)
                    elif len(column_names) > 1:
                        if '数量' not in column_names:
                            if '出库数量' in column_names:
                                df.rename(columns={'出库数量': '数量'}, inplace=True)
                                df.to_excel(file_path, index=False)
                            else:
                                print(column_names)
                                print(file_path)
                                print('-' * 200)
                                floder_name = data_path + '处理失败' + '\\' + '数量' + '\\' + '\\'.join(
                                    file_path.split('\\')[len(data_path.split('\\')):-1])
                                if not os.path.exists(floder_name):
                                    os.makedirs(floder_name)
                                new_file_path = floder_name + '\\' + fname
                                shutil.move(file_path, new_file_path)
                else:
                    print(f'空文件-> {file_path}')


# 深圳瑞霖：产品名称
def get_product_name(data_path):
    print('xy： 正在提取 产品名称 ,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    product_name_list = []
    with open(r'.\features\shxy\product_name.txt') as f:
        product_name_feature = f.readlines()[0]
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    if '产品名称' not in list(df.columns):
                        # docs = get_docs(df, product_name_feature)
                        df_object = CleanModel(file_path, features=product_name_feature)
                        columns_index = df_object.model()[0]
                        cos_max_sim = df_object.model()[1]
                        df = df_object.df
                        # tfidf_value = tfidf(df,product_name_feature)
                        # product_name_index, cos_sim_max = column_index(tfidf_value)
                        # product_name_list.append(cos_sim_max)
                        # df = pd.read_excel(file_path, dtype='object')
                        if cos_max_sim >= 0.01:  # cos_sim均值0.1645
                            df['产品名称'] = df.iloc[:, columns_index]
                            df.to_excel(file_path, index=False)
                        else:
                            floder_name = data_path + '处理失败' + '\\' + '产品名称' + '\\' + '\\'.join(
                                file_path.split('\\')[len(data_path.split('\\')):-1])
                            if not os.path.exists(floder_name):
                                os.makedirs(floder_name)
                            new_file_path = floder_name + '\\' + fname
                            shutil.move(file_path, new_file_path)
                            print(f'产品名称失败->|{new_file_path}')
                            print('-' * 200)
                            # print(cos_sim_max)
                            # print(list(df.columns)[product_name_index])
                            # print('=' * 130)


# 深圳瑞霖：规格
def get_product_spe(data_path):
    print('xy： 正在提取 规格,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    product_spe_list = []
    with open(r'.\features\shxy\product_spec.txt') as f:
        product_spe = f.readlines()[0]
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    if '规格' not in list(df.columns):
                        # docs = get_docs(df, product_spe)
                        tfidf_value = tfidf(df, product_spe)
                        product_spe_index, cos_sim_max = column_index(tfidf_value)
                        product_spe_list.append(cos_sim_max)
                        df = pd.read_excel(file_path, dtype='object')
                        if cos_sim_max >= 0.04:  # cos_sim均值0.1645
                            df['规格'] = df.iloc[:, product_spe_index]
                            df.to_excel(file_path, index=False)
                        else:
                            floder_name = data_path + '处理失败' + '\\' + '规格' + '\\' + '\\'.join(
                                file_path.split('\\')[len(data_path.split('\\')):-1])
                            if not os.path.exists(floder_name):
                                os.makedirs(floder_name)
                            new_file_path = floder_name + '\\' + fname
                            shutil.move(file_path, new_file_path)
                            print(f'xy： 规格 失败->|{new_file_path}')
                            print('-' * 200)
                            # print(file_path)
                            # print(cos_sim_max)
                            # print(list(df.columns)[product_spe_index])
                            # print('=' * 130)


# 深圳瑞霖：生产厂家
def get_product_manu(data_path):
    print('xy： 正在提取 生产厂家,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    with open(r'.\features\shxy\factory_manu.txt') as f:
        product_manu = f.readlines()[0]
    factory_name_list = []
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    if '生产厂家' not in list(df.columns):
                        # docs = get_docs(df, product_manu)
                        tfidf_value = tfidf(df, product_manu)
                        product_manu_index, cos_sim_max = column_index(tfidf_value)
                        factory_name_list.append(cos_sim_max)
                        df = pd.read_excel(file_path, dtype='object')
                        if cos_sim_max >= 0.04:  # cos_sim均值0.1645
                            df['生产厂家'] = df.iloc[:, product_manu_index]
                            df.to_excel(file_path, index=False)
                        else:
                            # print(file_path)
                            # print(cos_sim_max)
                            df['生产厂家'] = 'xy'
                            df.to_excel(file_path, index=False)
                            # print(file_path)
                            # print(cos_sim_max)
                            # print(list(df.columns)[product_manu_index])
                            # print('=' * 130)
                    # else:
                    # if len(df['生产厂家'].to_list())<df.shape[0]*0.9:
                    # df['生产厂家']='xy'
                    # df.to_excel(file_path, index=False)


# 深圳瑞霖：批号
def get_batch_num(data_path):
    print('xy：正在提取 批号 ,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    batch_num_list = []
    with open(r'.\features\shxy\batch_num.txt') as f:
        batch_num = f.readlines()[0]
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    if '批号' not in list(df.columns):
                        # docs = get_docs(df, batch_num)
                        tfidf_value = tfidf(df, batch_num)
                        batch_num_index, cos_sim_max = column_index(tfidf_value)
                        batch_num_list.append(cos_sim_max)
                        df = pd.read_excel(file_path, dtype='object')
                        if cos_sim_max >= 0.04:  # cos_sim均值0.1645
                            df['批号'] = df.iloc[:, batch_num_index]
                            df.to_excel(file_path, index=False)
                        else:

                            # floder_name = data_path + '处理失败' + '\\' + '批号' + '\\' + '\\'.join(
                            #     file_path.split('\\')[len(data_path.split('\\')):-1])
                            # if not os.path.exists(floder_name):
                            #     os.makedirs(floder_name)
                            # new_file_path = floder_name + '\\' + fname
                            # shutil.move(file_path, new_file_path)

                            print(f'xy： 批号 失败->|{file_path}')
                            print('-' * 200)
                            # print(file_path)
                            # print(cos_sim_max)
                            # print(list(df.columns)[batch_num_index])
                            # print('=' * 130)


# 深圳瑞霖：产品单位
def get_product_unit(data_path):
    print('xy： 正在提取 产品单位,请稍后...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                product_unit_count = 0
                list1 = []
                if df.shape[0] > 0:
                    for i in range(df.shape[1]):
                        for j in df.iloc[:, i]:
                            j = str(j).replace(' ', '')
                            if j == '支' or j == '瓶' or j == '盒' or j == '包' or j == '套' or j == 'KG' or j == '袋' or j == '听':
                                product_unit_count += 1
                        list1.append(product_unit_count)
                    if product_unit_count / df.shape[0] >= 0.5:
                        product_unit_index = list1.index(np.max(np.array(list1)))
                        df.rename(columns={list(df.columns)[product_unit_index]: '产品单位'}, inplace=True)
                        df.to_excel(file_path, index=False)
                    #             df['产品单位']=df.iloc[:,product_unit_index]
                    else:
                        df['产品单位'] = '缺失'
                        df.to_excel(file_path, index=False)


# 深圳瑞霖：单价
def get_product_price(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                for i in df.columns:
                    if '单价' in i or '零售价' in i or '售价' in i or i == '含税价':
                        price_index = list(df.columns).index(i)
                        df['单价'] = df.iloc[:, price_index]
                        df.to_excel(file_path, index=False)


# 深圳瑞霖：金额
def get_product_amount(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                for i in df.columns:
                    if '金额' in i:
                        amount_index = list(df.columns).index(i)
                        df['金额'] = df.iloc[:, amount_index]
                        df.to_excel(file_path, index=False)


# 深圳瑞霖：送货地址
def get_address(data_path):
    new_data_path = data_path + '处理成功'
    list2 = []
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    for i in df.columns:
                        if str(i).endswith('地址') or i == '发往地':
                            address_index = list(df.columns).index(i)
                            df['送货地址'] = df.iloc[:, address_index]
                            df.to_excel(file_path, index=False)


# 时间格式转化
def T_date(df):
    for i in df.columns:
        if df[i].dtype == 'datetime64[ns]':
            df[i] = df[i].apply(lambda x: str(pd.to_datetime(x).date()))


def check_key(data_path):
    print('正在检查 关键字段是否缺失...')
    print('=' * 200)
    new_data_path = data_path + '处理成功'
    list1 = ['销售日期', '客户名称', '产品名称', '规格', '产品单位', '数量', '生产厂家']
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path, dtype='object')
                if df.shape[0] > 0:
                    for i in list1:
                        if i not in list(df.columns):
                            new_floder = data_path + '处理失败' + '\\' + str(i) + '\\' + '\\'.join(
                                file_path.split('\\')[-2:-1])
                            if not os.path.exists(new_floder):
                                os.makedirs(new_floder)
                            new_file_path = os.path.join(new_floder, fname)
                            shutil.move(file_path, new_file_path)
                            print(f'{i} 字段缺失->| {file_path}')
                            print('-' * 200)
                else:
                    floder_name = data_path + '处理失败' + '\\' + '数据缺失' + '\\' + '\\'.join(
                        file_path.split('\\')[len(data_path.split('\\')):-1])
                    if not os.path.exists(floder_name):
                        os.makedirs(floder_name)
                    new_file_path = floder_name + '\\' + fname
                    shutil.move(file_path, new_file_path)

    print('xy：关键字段检查完成，准备清洗数据...')
    print('=' * 200)
    # print('=' * 200)
    #     for i in list1:
    #         if i not in list(df.columns):
    #             msg+=1
    #             print(f'{i} 字段缺失->| {file_path}')
    #             print('-'*200)
    # return msg


# 去除非必填字段，选填字段
def reduce_data(data_path):
    new_data_path = data_path + '处理成功'
    df_final = pd.DataFrame()
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[5:8] == 'SAL':
                # print(file_path)
                df = pd.read_excel(file_path, dtype='object')
                list1 = ['销售日期', '客户名称', '产品名称', '规格', '产品单位', '数量', '批号', '单价', '金额', '送货地址', '生产厂家', '产品编码', '客户代码',
                         '效期']
                list2 = list(df.columns)
                column_names = [column_name for column_name in list1 if column_name in list2]
                df_final = df[column_names]
                df.dropna(how='any', axis=0, inplace=True)
                # df.to_excel(file_path, index=False)
                df_final = df_final[(df_final['数量'] != 0) & (df_final['数量'] != '0') & (df_final['数量'] != '&nbsp;') \
                                    & (df_final['数量'].notnull()) & (df_final['数量'] != '合计：')]
                T_date(df_final)
                df_final.to_excel(file_path, index=False)


def del_flows(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[5:8] == 'SAL':
                # print(file_path)
                df = pd.read_excel(file_path, dtype='object')
                df = df[
                    (df['销售日期'].notnull()) & (df['销售日期'] != '合计') & (df['销售日期'] != '合计：') & (df['销售日期'] != 'NaT') & (
                                df['销售日期'] != '业务日期')]
                df = df[(df['产品单位'].notnull()) & (df['产品单位'] != '----------')]
                df = df[(df['客户名称'].notnull()) & (df['客户名称'] != '~')]
                df.to_excel(file_path, index=False)


def sale_date_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            # print(file_path)
            df = pd.read_excel(file_path, dtype='object')
            T_date(df)
            if fname[5:8] == 'SAL' and df.shape[0] > 0:
                list3 = []
                for i in df['销售日期']:
                    try:
                        i = pd.to_datetime(str(i)[:10]).strftime('%Y-%m-%d')
                        list3.append(i)
                    except:
                        try:
                            # print(file_path)
                            i = pd.to_datetime(str(i)[:9]).strftime('%Y-%m-%d')
                            list3.append(i)
                        except:
                            floder_name = data_path + '处理失败' + '\\' + '销售日期' + '\\' + '\\'.join(
                                file_path.split('\\')[len(data_path.split('\\')):-1])
                            if not os.path.exists(floder_name):
                                os.makedirs(floder_name)
                            new_file_path = floder_name + '\\' + fname
                            try:
                                shutil.move(file_path, new_file_path)
                                print(f'销售日期 清洗失败->|{new_file_path}')
                                print('-' * 200)
                            except:
                                pass
                    try:
                        df['销售日期'] = list3
                        df.to_excel(file_path, index=False)
                    except:
                        pass
                    # try:
                    #     df['销售日期'] = list3
                    #     df.to_excel(file_path, index=False)
                    # except:
                    #     floder_name = data_path + '处理失败' + '\\' + '销售日期' + '\\' + '\\'.join(
                    #         file_path.split('\\')[len(data_path.split('\\')):-1])
                    #     if not os.path.exists(floder_name):
                    #         os.makedirs(floder_name)
                    #     new_file_path = floder_name + '\\' + fname
                    #     try:
                    #         shutil.move(file_path, new_file_path)
                    #         print(f'销售日期 清洗失败->|{new_file_path}')
                    #         print('-' * 200)
                    #     except:
                    #         pass


def shxy_excel_style(df, data_path, fname):
    writer = pd.ExcelWriter(os.path.join(data_path, fname), engine='openpyxl')
    df.to_excel(writer, index=False)
    worksheet = writer.sheets['Sheet1']
    font = Font(name='微软雅黑', bold=True, color='f7f7f7')
    alignment = Alignment(vertical='top', wrap_text=True)
    pattern_fill = PatternFill(fill_type='solid', fgColor='00b0f0')
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in itertools.chain(*worksheet['A1:O1']):
        cell.font = font
        cell.alignment = alignment
        cell.fill = pattern_fill
        cell.border = border
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 8
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 8
    worksheet.column_dimensions['K'].width = 8
    worksheet.column_dimensions['L'].width = 35
    worksheet.column_dimensions['M'].width = 30
    worksheet.column_dimensions['N'].width = 12
    worksheet.column_dimensions['O'].width = 35

    writer.save()
    writer.close()


def check_data(data_path):
    new_data_path = data_path + '处理结果'
    concat_df = pd.DataFrame()
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[5:8] == 'SAL':
                df = pd.read_excel(file_path, dtype='object')
                df['经销商'] = file_path.split('\\')[-2].split('_')[1]
                concat_df = pd.concat([concat_df, df], axis=0)
    list1 = ['销售日期', '客户名称', '产品名称', '规格', '产品单位', '数量', '批号', '单价', '金额', '送货地址', '生产厂家', '产品编码', '客户代码', '效期', '经销商']
    list2 = list(concat_df.columns)
    column_name = [x for x in list1 if x not in list2]
    for i in column_name:
        concat_df[i] = ''
    concat_df = concat_df[
        ['销售日期', '客户代码', '客户名称', '产品编码', '产品名称', '规格', '产品单位', '数量', '批号', '单价', '金额', '送货地址', '生产厂家', '效期', '经销商']]
    # shxy_excel_style(concat_df,'\\'.join(data_path.split('\\')[:-1]),data_path.split('\\')[-1]+'数据合并SAL.xlsx')
    concat_df.to_excel(data_path + '数据合并SAL.xlsx', index=False)


def add_factory_manue(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            file_path = os.path.join(dirpath, fname)
            if fname[5:8] == 'SAL':
                df = pd.read_excel(file_path, dtype='object')
                if '生产厂家' not in list(df.columns):
                    df['生产厂家'] = 'xy'  # df['生产厂家'].fillna('xy')
                    df.to_excel(file_path, index=False)


def shxy_control_program():
    # data_path=r'C:\Users\guodingyu\Downloads\接收文件一级商'
    # copy_path = r'C:\Users\guodingyu\Downloads\备份'  # 原始文件备份文件夹
    data_path = r'.\data\接收文件'
    copy_path = r'.\data\数据备份'  # 原始文件备份文件夹
    receive_date = ''.join(str(datetime.datetime.now())[:10].split(sep='-'))
    receive_data = data_path + '\\' + receive_date + '接收文件'
    now_path = os.path.join(data_path, receive_date)  # 当前处理文件夹，即需要监控的文件夹
    rusult_data = data_path + '\\' + receive_date + '处理结果'
    if not os.path.exists(now_path):
        os.makedirs(now_path)  # 创建处理文件夹
    if not os.path.exists(receive_data):
        os.makedirs(receive_data)  # 创建日累计接收文件夹
    if not os.path.exists(rusult_data):
        os.makedirs(rusult_data)  # 创建处理结果文件夹
    floder_list = []
    for floder in os.listdir(now_path):
        floder_list.append(floder)
    msg = str()
    if len(floder_list) == 0:
        msg = 'N'
        print('=' * 74 + f' {str(datetime.datetime.now())[:19]} 数据分拣系统 实时监控 无数据  ' + '=' * 73)
        print('-' * 200)
        time.sleep(180)
        shxy_control_program()
    else:
        msg = 'Y'
        print(f'{str(datetime.datetime.now())[:19]} 数据分拣系统 实时监控 来数据了...')
        print('=' * 200)
        for floder in os.listdir(now_path):
            now_floder = os.path.join(now_path, floder)
            new_receive_data = os.path.join(receive_data, floder)
            new_copy_path = os.path.join(copy_path + '\\' + receive_date, floder)
            if not os.path.exists(new_receive_data):
                shutil.copytree(now_floder, new_receive_data)  # 备份每日原始文件夹
            if not os.path.exists(new_copy_path):
                shutil.copytree(now_floder, new_copy_path)  # 备份项目原始文件夹
    return msg, now_path, rusult_data


# 批号数据清洗
def sal_batch_num_clean(data_path):
    new_data_path = data_path + '处理成功'
    for dirpath, dirname, filenames in os.walk(new_data_path):
        for fname in filenames:
            if fname[5:8] == 'SAL':
                file_path = os.path.join(dirpath, fname)
                df = pd.read_excel(file_path)
                if df.shape[0] > 1 and '批号' in list(df.columns):
                    df['批号'] = df['批号'].map(lambda x: str(x))
                    df['批号'] = df['批号'].str.extract('([a-zA-Z]\d+|\d+)')
                    df = df.replace('nan', '')
                    df.to_excel(file_path, index=False)


def sal_step1(data_path):
    time_start = time.time()  # 记录开始时间
    # if not os.path.exists(data_path+'处理成功'):
    shxy_sta(data_path)  # 标准化  经销商识别
    get_spe_sal(data_path)  # 黑名单经销商
    del_temporary(data_path)  # 删除临时文件
    file_conversion(data_path)  # 文件格式转换
    check_conversion(data_path)  # 检查转换是否成功
    cheack_file(data_path)  # 检查转换是否成功  检查后缀名
    del_empty_floder(data_path)  # 删除空文件夹
    shxy_spe.shxy_spe(data_path)  # 特殊经销商清洗
    rename_file(data_path)  # 重命名文件
    cheack_sheet(data_path)  # 获取sheet检查
    get_sheets(data_path)  # 多sheet识别  移动文件
    sheets_split(data_path)  # 多sheet拆分
    empty_file(data_path)  # 空文件识别  移动文件，即 经销商
    get_header(data_path)  # 确定表头
    split_columns(data_path)  # 品名-规格-产地 分列
    sale_date_clean1(data_path)  # 销售日期修改
    del_nullstr(data_path)  # 删除表头空字符串
    del_columns(data_path)  # 删除容干扰的空列
    batch_num_clean(data_path)  # 批号修改
    product_name_clean(data_path)  # 产品名称修改
    product_spe_clean(data_path)  # 规格修改
    custer_name_clean(data_path)  # 客户名称修改
    product_num_clean(data_path)  # 数量
    custer_name_id(data_path)  # 客户编码
    product_name_id(data_path)  # 产品编码
    deter_date_clean(data_path)  # 效期
    get_deter_date(data_path)  # 效期
    time_end = time.time()
    time_sum = time_end - time_start
    print(f'shxy_sal_step1 running->| {round(time_sum, 2)}s ')
    print('=' * 200)


def sal_step2(data_path):
    time_start = time.time()  # 记录开始时间
    get_custer_name(data_path)  # 客户名称
    get_product_num(data_path)  # 数量
    get_sale_date(data_path)  # 销售日期
    get_product_name(data_path)  # 产品名称
    get_product_spe(data_path)  # 规格
    get_product_manu(data_path)  # 生产厂家
    get_batch_num(data_path)  # 批号
    get_product_unit(data_path)  # 产品单位
    get_product_price(data_path)  # 单价
    get_product_amount(data_path)  # 金额
    get_address(data_path)  # 送货地址
    time_end = time.time()
    time_sum = time_end - time_start
    print(f'shxy_sal_step2 running->| {round(time_sum, 2)}s ')
    print('=' * 200)


def sal_step3(data_path):
    drop_msg = str()
    try:
        time_start = time.time()  # 记录开始时间
        check_key(data_path)  # 检查必填字段是否缺失
        reduce_data(data_path)
        del_flows(data_path)
        sale_date_clean(data_path)
        add_factory_manue(data_path)  # 生产厂家缺失的补
        sal_batch_num_clean(data_path)  # 批号数据清洗
        # check_data(data_path)
        time_end = time.time()
        time_sum = time_end - time_start
        print(f'shxy_sal_step3 running->| {round(time_sum, 2)}s ')
        print('=' * 200)
        drop_msg = 'Y'
    except:
        drop_msg = 'N'
    return drop_msg


if __name__ == '__main__':
    while 1 == 1:
        msg, data_path, rusult_data = shxy_control_program()
        if msg == 'Y':
            sal_step1(data_path)
            sal_step2(data_path)
            shxy_pur.pur_clean(data_path)
            shxy_inv.inv_clean(data_path)
            drop_msg = sal_step3(data_path)
            report_data_path = '\\'.join(data_path.split('\\')[:-1])
            report_obj = ReportCleanData(data_path=report_data_path)
            report = report_obj.makereport()
            if drop_msg == 'Y':
                for floder in os.listdir(data_path):  # 清空原始文件夹
                    floder_name = os.path.join(data_path, floder)
                    # print(floder_name)
                    shutil.rmtree(floder_name)
                now_data = data_path + '处理成功'
                for rusult_floder in os.listdir(now_data):  # 清空处理完成文件夹，放置处理结果文件夹
                    floder_path = os.path.join(now_data, rusult_floder)
                    new_floder_path = os.path.join(rusult_data, rusult_floder)
                    shutil.move(floder_path, new_floder_path)
            check_data(data_path)
